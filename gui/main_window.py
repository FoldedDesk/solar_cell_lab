"""
主窗口 — 虚拟实验台
实验一：固定光源+距离(30cm)，改变电阻 → 用户手动填 U, I
实验二：固定光源+电阻，改变距离 → 记录 d, I, Voc, Isc
实验三：固定距离+电阻，改变光强 → 记录 E, U, I, P
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import numpy as np
import matplotlib
matplotlib.use("TkAgg")
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import matplotlib.font_manager as fm
from scipy.interpolate import PchipInterpolator
from PIL import Image, ImageTk

# macOS/Windows 中文兼容字体回退
plt_font = fm.FontProperties(family=[
    "PingFang SC", "Hiragino Sans GB", "Microsoft YaHei", "STHeiti", "SimHei"
])

import sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from physics import (solve_iv_curve, solve_voc, solve_isc, find_mpp,
                     calc_fill_factor, DEFAULT_PARAMS, calc_light_intensity,
                     get_experimental_data)

# ── 颜色主题 ──
BG = "#d7dadd"
PANEL_BG = "#d5e7f2"
FG = "#1f2d3d"
ACCENT = "#0086c8"
ACCENT2 = "#1f9f59"
GRID_COLOR = "#8da0b1"

def configure_ttk_theme(root):
    """统一 ttk 在 Windows/macOS 的可读性，避免 mac 下白底浅字。"""
    style = ttk.Style(root)
    # Windows 的 vista/xpnative 主题常忽略按钮背景，但会吃前景色，容易出现白底白字。
    # 统一使用 clam，确保跨平台颜色样式一致可控。
    try:
        style.theme_use("clam")
    except Exception:
        pass

    style.configure(".", background=BG, foreground=FG)
    style.configure("TFrame", background=BG)
    style.configure("TLabel", background=BG, foreground=FG)
    style.configure("TNotebook", background=BG, borderwidth=0)
    style.configure("TNotebook.Tab", background=PANEL_BG, foreground=FG, padding=(12, 6))
    style.map(
        "TNotebook.Tab",
        background=[("selected", "#e9f4fa"), ("active", "#dcecf5")],
        foreground=[("selected", "#102235"), ("active", "#1b3b5a")],
    )
    style.configure("Treeview", background="#f2f4f7", fieldbackground="#f2f4f7", foreground="#111")
    style.configure("Treeview.Heading", background=PANEL_BG, foreground="#12283d")
    style.map("Treeview", background=[("selected", "#cddff0")], foreground=[("selected", "#111")])
    # 彩色按钮样式（跨平台可控）
    style.configure("Primary.TButton", background="#2a6", foreground="#ffffff", padding=(10, 6), borderwidth=1)
    style.map(
        "Primary.TButton",
        background=[("active", "#3b7"), ("disabled", "#b9d8c7")],
        foreground=[("disabled", "#5f6b63"), ("!disabled", "#ffffff")],
    )
    style.configure("Info.TButton", background="#2266cc", foreground="#ffffff", padding=(10, 6), borderwidth=1)
    style.map(
        "Info.TButton",
        background=[("active", "#3377dd"), ("disabled", "#bfd0ec")],
        foreground=[("disabled", "#5a6472"), ("!disabled", "#ffffff")],
    )
    style.configure("Assist.TButton", background="#4455aa", foreground="#ffffff", padding=(10, 6), borderwidth=1)
    style.map(
        "Assist.TButton",
        background=[("active", "#5566bb"), ("disabled", "#c6cce4")],
        foreground=[("disabled", "#5a6070"), ("!disabled", "#ffffff")],
    )


def solve_operating_point(params, R_ohm):
    """给定负载电阻 R，求工作点 (V, I, P)。"""
    if R_ohm <= 0:
        return 0.0, 0.0, 0.0
    V_arr, I_arr, _ = solve_iv_curve(params)
    I_load = V_arr / R_ohm
    diff = I_arr - I_load
    for i in range(len(diff) - 1):
        if diff[i] * diff[i + 1] <= 0:
            t = diff[i] / (diff[i] - diff[i + 1])
            V_op = V_arr[i] + t * (V_arr[i + 1] - V_arr[i])
            I_op = I_arr[i] + t * (I_arr[i + 1] - I_arr[i])
            return float(V_op), float(I_op), float(V_op * I_op)
    idx = np.argmin(np.abs(diff))
    return float(V_arr[idx]), float(I_arr[idx]), float(V_arr[idx] * I_arr[idx])


# ═══════════════════════════════════════════════
#  手动输入对话框
# ═══════════════════════════════════════════════

class _ManualInputDialog:
    """弹窗让用户手动输入 U (V) 和 I (mA)。
    Enter 在 U 框 → 跳到 I 框；Enter 在 I 框 → 确认提交。
    """

    def __init__(self, parent, R):
        self.result = None
        self.top = tk.Toplevel(parent)
        self.top.title("手动输入测量数据")
        self.top.configure(bg=BG)
        self.top.geometry("360x240")
        self.top.resizable(False, False)
        self.top.grab_set()

        tk.Label(self.top, text="记录数据点", bg=ACCENT, fg="#fff",
                 font=("Microsoft YaHei", 13, "bold")).pack(fill=tk.X, ipady=6)

        tk.Label(self.top, text="当前电阻 R = {:g} Ω".format(R),
                 bg=BG, fg="#aaa", font=("Microsoft YaHei", 10)).pack(pady=(10, 12))

        # U 输入
        f1 = tk.Frame(self.top, bg=BG)
        f1.pack(fill=tk.X, padx=30, pady=4)
        tk.Label(f1, text="电压 U (V):", bg=BG, fg=FG,
                 font=("Microsoft YaHei", 11), width=12, anchor="e").pack(side=tk.LEFT)
        self.u_var = tk.StringVar(value="")
        self.u_entry = tk.Entry(f1, textvariable=self.u_var, font=("Consolas", 13),
                                bg="#1a1a1a", fg="#00ff88", insertbackground="#00ff88",
                                width=12, relief=tk.SUNKEN, bd=2)
        self.u_entry.pack(side=tk.LEFT, padx=8)
        self.u_entry.focus_set()

        # I 输入
        f2 = tk.Frame(self.top, bg=BG)
        f2.pack(fill=tk.X, padx=30, pady=4)
        tk.Label(f2, text="电流 I (mA):", bg=BG, fg=FG,
                 font=("Microsoft YaHei", 11), width=12, anchor="e").pack(side=tk.LEFT)
        self.i_var = tk.StringVar(value="")
        self.i_entry = tk.Entry(f2, textvariable=self.i_var, font=("Consolas", 13),
                                bg="#1a1a1a", fg="#00ff88", insertbackground="#00ff88",
                                width=12, relief=tk.SUNKEN, bd=2)
        self.i_entry.pack(side=tk.LEFT, padx=8)

        # Enter 键绑定
        self.u_entry.bind("<Return>", lambda e: self.i_entry.focus_set())
        self.i_entry.bind("<Return>", lambda e: self._ok())

        # 按钮
        bf = tk.Frame(self.top, bg=BG)
        bf.pack(fill=tk.X, padx=30, pady=(12, 0))
        ttk.Button(bf, text="确认", style="Primary.TButton",
                   command=self._ok).pack(side=tk.LEFT, expand=True, padx=4)
        ttk.Button(bf, text="取消", style="Assist.TButton",
                   command=self._cancel).pack(side=tk.LEFT, expand=True, padx=4)

        self.top.bind("<Escape>", lambda e: self._cancel())

    def _ok(self):
        try:
            u = float(self.u_var.get().strip())
            i = float(self.i_var.get().strip())
        except ValueError:
            self._flash_error("请输入有效数字！")
            return
        if u < 0 or i < 0:
            self._flash_error("数值不能为负！")
            return
        self.result = (u, i)
        self.top.destroy()

    def _cancel(self):
        self.result = None
        self.top.destroy()

    def _flash_error(self, msg):
        for w in self.top.winfo_children():
            if isinstance(w, tk.Label) and w.cget("bg") == "#ff4444":
                w.destroy()
        tk.Label(self.top, text=msg, bg="#ff4444", fg="#fff",
                 font=("Microsoft YaHei", 10, "bold")).pack(fill=tk.X)


# ═══════════════════════════════════════════════
#  实验一：改变电阻（手动填数据）
# ═══════════════════════════════════════════════

class ExperimentOneTab:
    """实验一：固定光源+距离，改变电阻，用户手动输入 U、I。"""

    def __init__(self, parent_frame, experiment_name="实验一"):
        self.frame = parent_frame
        self.experiment_name = experiment_name
        self.is_distance_experiment = (self.experiment_name == "实验二")
        self.title = "改变负载电阻"
        self.params = {k: v for k, v in DEFAULT_PARAMS["single"].items()}
        self.data_points = []
        self.distance_cm = 30.0
        self.light_intensity = 105.1 if self.is_distance_experiment else 242.0
        self.panel_area_m2 = 0.01  # 受光面积(约 10cm x 10cm)
        self.wires = set()
        self.term_pos = {}
        self.term_degree = {}
        self.drag_start_term = None
        self.drag_line_id = None
        self.drag_device_id = None
        self.drag_device_last_xy = None
        self.connection_ok = False
        self.last_auto_record_r = None
        self.r_value = 82.0 if self.is_distance_experiment else 0.0
        self.distance_resistance_confirmed = False
        self.exp2_measure_mode = "isc"  # 实验二：isc(短路电流) / voc(开路电压)
        self.exp2_partial_points = {}
        self.last_auto_record_distance = None
        self.distance_win = None
        self.distance_scale = None
        self.distance_preview_canvas = None
        self.distance_indicator_id = None
        self.lab_win = None
        self.circuit_canvas = None
        self.table_win = None
        self.curve_win = None
        self.tree = None
        self.fig1 = None
        self.fig2 = None
        self.ax1 = None
        self.ax2 = None
        self.canvas1 = None
        self.canvas2 = None
        self._asset_src = {}
        self._asset_cache = {}
        self._asset_tk_refs = {}
        self.curve_fit_mode_1 = False
        self.curve_fit_mode_2 = False
        self.curve_btn1 = None
        self.curve_btn2 = None
        self.scene_size = (250, 320)
        self.devices = {
            "lamp": {"x": 72, "y": 86, "w": 104, "h": 52, "title": "大功率灯", "fill": "#3a2f1b"},
            "panel": {"x": 404, "y": 72, "w": 122, "h": 78, "title": "太阳能板", "fill": "#17324a"},
            "amm": {"x": 96, "y": 282, "w": 96, "h": 96, "title": "电流表", "fill": "#1f3f2b"},
            "res": {"x": 478, "y": 282, "w": 96, "h": 96, "title": "电阻箱", "fill": "#40291f"},
            "vol": {"x": 284, "y": 456, "w": 96, "h": 96, "title": "电压表", "fill": "#33295a"},
        }
        self.device_term_offsets = {
            "panel": {"panel_p": (0, 22), "panel_n": (122, 48)},
        }
        self.device_icon_bounds = {}
        self._load_scene_assets()
        self._build()

    def _load_scene_assets(self):
        base = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "res")
        mapping = {
            "amm": os.path.join(base, "ammeter_256.png"),
            "vol": os.path.join(base, "voltmeter_256.png"),
            "res": os.path.join(base, "resistance_box_256.png"),
        }
        for k, p in mapping.items():
            if os.path.exists(p):
                try:
                    self._asset_src[k] = Image.open(p).convert("RGBA")
                except Exception:
                    pass

    def _get_device_asset(self, key, w, h):
        if key not in self._asset_src:
            return None
        size_key = (key, int(w), int(h))
        if size_key in self._asset_tk_refs:
            return self._asset_tk_refs[size_key]
        src = self._asset_src[key]
        tw, th = int(w), int(h)
        sw, sh = src.size
        scale = min(tw / max(sw, 1), th / max(sh, 1))
        nw = max(1, int(sw * scale))
        nh = max(1, int(sh * scale))
        fitted = src.resize((nw, nh), Image.Resampling.LANCZOS)
        # 等比缩放后居中贴到目标画布，避免任何方向拉伸
        img = Image.new("RGBA", (tw, th), (0, 0, 0, 0))
        ox = (tw - nw) // 2
        oy = (th - nh) // 2
        img.paste(fitted, (ox, oy), fitted)
        tk_img = ImageTk.PhotoImage(img)
        self._asset_tk_refs[size_key] = tk_img
        return tk_img

    def _build(self):
        body = tk.Frame(self.frame, bg=BG)
        body.pack(fill=tk.BOTH, expand=True)
        mid = tk.Frame(body, bg=BG)
        mid.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 6))
        right = tk.Frame(body, bg=PANEL_BG, width=520)
        right.pack(side=tk.LEFT, fill=tk.Y)
        right.pack_propagate(False)
        self._build_wiring_scene(mid, large=True, show_button=False)
        self._build_right(right)

    # ── 左侧控制面板 ──

    def _build_control(self, parent):
        # 电池类型（仅单晶硅，无选择）
        tk.Label(parent, text="电池类型", bg=PANEL_BG, fg=FG,
                 font=("Microsoft YaHei", 10, "bold")).pack(anchor="w", padx=10, pady=(10, 2))
        tf = tk.Frame(parent, bg=PANEL_BG)
        tf.pack(fill=tk.X, padx=10)
        self.cell_type_var = tk.StringVar(value="single")
        tk.Label(tf, text="单晶硅", bg=PANEL_BG, fg=FG,
                 font=("Microsoft YaHei", 9)).pack(side=tk.LEFT, padx=5)
        ttk.Separator(parent, orient="horizontal").pack(fill=tk.X, padx=10, pady=6)

        # 当前电阻值（由实验台双击电阻箱调整）
        tk.Label(parent, text="━━━ 负载电阻 ━━━", bg=PANEL_BG, fg=ACCENT,
                 font=("Microsoft YaHei", 10, "bold")).pack(anchor="w", padx=10, pady=(0, 2))
        self.r_readonly_label = tk.Label(parent, text="R = 0.0 Ω（双击实验台电阻箱修改）",
                                         bg="#1e1e1e", fg="#00ff88",
                                         font=("Consolas", 11, "bold"),
                                         relief=tk.SUNKEN, bd=1, padx=8, pady=6)
        self.r_readonly_label.pack(fill=tk.X, padx=10, pady=(2, 6))

        # 固定条件
        info = tk.Frame(parent, bg=PANEL_BG)
        info.pack(fill=tk.X, padx=10, pady=4)
        tk.Label(info, text="固定条件:", bg=PANEL_BG, fg="#888",
                 font=("Microsoft YaHei", 9)).pack(anchor="w")
        if self.is_distance_experiment:
            tk.Label(info, text="  距离 d: 可调（双击太阳能板）", bg=PANEL_BG, fg=FG,
                     font=("Microsoft YaHei", 9)).pack(anchor="w")
            tk.Label(info, text="  初始光强 I: {:.1f} W/m²".format(self.light_intensity), bg=PANEL_BG, fg=FG,
                     font=("Microsoft YaHei", 9)).pack(anchor="w")
            tk.Label(info, text="  提示: 每个距离需分两次接线测 Isc 和 Voc", bg=PANEL_BG, fg="#8a4b08",
                     font=("Microsoft YaHei", 9, "bold")).pack(anchor="w")
        else:
            tk.Label(info, text="  光强 I: 242 W/m²", bg=PANEL_BG, fg=FG,
                     font=("Microsoft YaHei", 9)).pack(anchor="w")
        tk.Label(info, text="  受光面积 A: 0.010 m²", bg=PANEL_BG, fg=FG,
                 font=("Microsoft YaHei", 9)).pack(anchor="w")

        core = tk.Frame(parent, bg="#1a2238", bd=1, relief=tk.SOLID)
        core.pack(fill=tk.X, padx=10, pady=(4, 2))
        tk.Label(core, text="核心原理：光功率 -> 电功率", bg="#24345f", fg="#dfe9ff",
                 font=("Microsoft YaHei", 9, "bold")).pack(fill=tk.X, ipady=2)
        self.core_pin_lbl = tk.Label(core, text="入射光功率 Pin = -- W", bg="#1a2238", fg="#9fc3ff",
                                     font=("Consolas", 10, "bold"), anchor="w")
        self.core_pin_lbl.pack(fill=tk.X, padx=6, pady=(4, 1))
        self.core_pout_lbl = tk.Label(core, text="输出电功率 Pout = -- W", bg="#1a2238", fg="#8dffb7",
                                      font=("Consolas", 10, "bold"), anchor="w")
        self.core_pout_lbl.pack(fill=tk.X, padx=6, pady=1)
        self.core_eta_lbl = tk.Label(core, text="转换效率 η = -- %", bg="#1a2238", fg="#ffd38a",
                                     font=("Consolas", 10, "bold"), anchor="w")
        self.core_eta_lbl.pack(fill=tk.X, padx=6, pady=(1, 5))
        self._update_core_principle(0.0, 0.0)

        ttk.Separator(parent, orient="horizontal").pack(fill=tk.X, padx=10, pady=6)

        if self.is_distance_experiment:
            mode_box = tk.Frame(parent, bg=PANEL_BG)
            mode_box.pack(fill=tk.X, padx=10, pady=(0, 6))
            tk.Label(mode_box, text="实验二测量模式", bg=PANEL_BG, fg=ACCENT,
                     font=("Microsoft YaHei", 10, "bold")).pack(anchor="w")
            mode_btns = tk.Frame(mode_box, bg=PANEL_BG)
            mode_btns.pack(fill=tk.X, pady=(2, 0))
            ttk.Button(mode_btns, text="测短路电流 Isc", style="Assist.TButton",
                       command=lambda: self._set_exp2_measure_mode("isc")).pack(side=tk.LEFT, expand=True, fill=tk.X, padx=(0, 4), ipady=1)
            ttk.Button(mode_btns, text="测开路电压 Voc", style="Assist.TButton",
                       command=lambda: self._set_exp2_measure_mode("voc")).pack(side=tk.LEFT, expand=True, fill=tk.X, padx=(4, 0), ipady=1)

        # 操作按钮
        record_btn_text = "【实验二】点击记录当前数据" if self.is_distance_experiment else "● 自动采集并记录"
        ttk.Button(parent, text=record_btn_text, style="Primary.TButton",
                   command=self._record_point).pack(fill=tk.X, padx=10, pady=4, ipady=2)

        btn_frame = tk.Frame(parent, bg=PANEL_BG)
        btn_frame.pack(fill=tk.X, padx=10, pady=4)

        ttk.Button(parent, text="数据分析", style="Info.TButton",
                   command=self._show_analysis).pack(fill=tk.X, padx=10, pady=6, ipady=2)

        ttk.Separator(parent, orient="horizontal").pack(fill=tk.X, padx=10, pady=6)

        ttk.Button(parent, text="打开实验台（独立页面）", style="Assist.TButton",
                   command=self._open_lab_window).pack(fill=tk.X, padx=10, pady=(2, 6), ipady=1)

    def _build_wiring_scene(self, parent, large=False, show_button=True):
        tk.Label(parent, text="━━━ 实验接线场景 ━━━", bg=BG, fg="#114a76",
                 font=("Microsoft YaHei", 11, "bold")).pack(anchor="w", padx=10, pady=(0, 2))
        self.wire_status = tk.Label(parent, text="接线状态: 未完成", bg=BG, fg="#9a4f00",
                                    font=("Microsoft YaHei", 10, "bold"))
        self.wire_status.pack(anchor="w", padx=10, pady=(0, 4))

        cw, ch = (960, 620) if large else (250, 320)
        self.scene_size = (cw, ch)
        self.scene = tk.Canvas(parent, width=cw, height=ch, bg="#d7dadd",
                               highlightthickness=1, highlightbackground="#2d3c5a")
        self.scene.pack(padx=10, pady=2)
        self.scene.bind("<ButtonPress-1>", self._on_scene_press)
        self.scene.bind("<B1-Motion>", self._on_scene_drag)
        self.scene.bind("<ButtonRelease-1>", self._on_scene_release)
        self.scene.bind("<Double-Button-1>", self._on_scene_double_click)
        if show_button:
            ttk.Button(parent, text="一键连线", style="Assist.TButton",
                       command=self._auto_wire_for_test).pack(fill=tk.X, padx=10, pady=(2, 4), ipady=1)
        timer_bar = tk.Frame(parent, bg=PANEL_BG)
        timer_bar.pack(fill=tk.X, padx=10, pady=(2, 6))
        tk.Label(timer_bar, text="实验已经进行:", bg=PANEL_BG, fg="#ba2f2f",
                 font=("Microsoft YaHei", 10, "bold")).pack(side=tk.LEFT)
        self.timer_var = tk.StringVar(value="00:00:00")
        tk.Label(timer_bar, textvariable=self.timer_var, bg="#101010", fg="#ff2a2a",
                 font=("Consolas", 16, "bold"), padx=8, pady=2).pack(side=tk.LEFT, padx=8)
        self._start_timer()
        self._draw_scene()

    def _start_timer(self):
        if not hasattr(self, "timer_var"):
            return
        if not hasattr(self, "_elapsed"):
            self._elapsed = 0
        h = self._elapsed // 3600
        m = (self._elapsed % 3600) // 60
        s = self._elapsed % 60
        self.timer_var.set(f"{h:02d}:{m:02d}:{s:02d}")
        self._elapsed += 1
        self.frame.after(1000, self._start_timer)

    def _open_lab_window(self):
        if self.lab_win is not None and self.lab_win.winfo_exists():
            self.lab_win.lift()
            self.lab_win.focus_set()
            return
        win = tk.Toplevel(self.frame.winfo_toplevel())
        win.title("{}独立实验台".format(self.experiment_name))
        win.configure(bg=BG)
        win.geometry("1080x700")
        win.resizable(True, True)
        self.lab_win = win
        self._layout_devices_for_large_scene()

        body = tk.Frame(win, bg=BG)
        body.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        left = tk.Frame(body, bg=PANEL_BG)
        left.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        right = tk.Frame(body, bg=PANEL_BG, width=270)
        right.pack(side=tk.LEFT, fill=tk.Y, padx=(8, 0))
        right.pack_propagate(False)

        self._build_wiring_scene(left, large=True, show_button=False)
        self._build_autowire_buttons(right, pady=(12, 6))
        tk.Label(right, text="右下角显示实时电路图", bg=PANEL_BG, fg="#a8bde6",
                 font=("Microsoft YaHei", 10, "bold")).pack(anchor="w", padx=10, pady=(8, 4))
        self.circuit_canvas = tk.Canvas(win, width=320, height=250, bg="#f7f4e8",
                                        highlightthickness=1, highlightbackground="#2d3340")
        self.circuit_canvas.place(relx=1.0, rely=1.0, x=-14, y=-14, anchor="se")
        self.circuit_canvas.tkraise()
        self._update_circuit_diagram()

        def _on_close():
            self.scene = None
            self.wire_status = None
            self.circuit_canvas = None
            self.lab_win = None
            win.destroy()

        win.protocol("WM_DELETE_WINDOW", _on_close)

    def _add_device(self, x, y, w, h, title, fill):
        self.scene.create_rectangle(x + 3, y + 3, x + w + 3, y + h + 3, fill="#0a0f1c", outline="", width=0, tags=("device",))
        self.scene.create_rectangle(x, y, x + w, y + h, fill=fill, outline="#5f79a8", width=1, tags=("device",))
        self.scene.create_rectangle(x + 1, y + 1, x + w - 1, y + 16, fill="#22365d", outline="", width=0, tags=("device",))
        self.scene.create_text(x + w / 2, y + 9, text=title, fill="#e8f1ff",
                               font=("Microsoft YaHei", 8, "bold"), tags=("device",))

    def _add_terminal(self, term_id, x, y, mark=""):
        r = 4
        self.term_pos[term_id] = (x, y)
        self.scene.create_oval(x - r, y - r, x + r, y + r, fill="#ffd166", outline="#ffe7a6",
                               width=1, tags=("terminal", "term_" + term_id))
        if mark:
            self.scene.create_text(x, y - 10, text=mark, fill="#9bb2dd", font=("Consolas", 8, "bold"))

    def _draw_scene(self):
        if self.scene is None or not self.scene.winfo_exists():
            return
        self.scene.delete("all")
        self.term_pos.clear()
        self.term_degree = {
            "panel_p": 0, "panel_n": 0,
            "amm_p": 0, "amm_n": 0,
            "res_p": 0, "res_n": 0,
            "vol_p": 0, "vol_n": 0
        }
        self.drag_start_term = None
        self.drag_line_id = None
        self.drag_device_id = None
        self.drag_device_last_xy = None
        self.device_icon_bounds.clear()

        cw, ch = self.scene_size
        # 顶部蓝色标题条（参考示例 UI）
        self.scene.create_rectangle(0, 0, cw, 26, fill="#0a79a9", outline="#075b7f", width=1)
        self.scene.create_rectangle(0, 0, cw, 6, fill="#3da8d1", outline="", width=0)
        self.scene.create_text(14, 13, text="太阳能电池的特性测量", anchor="w",
                               fill="#e9f7ff", font=("Microsoft YaHei", 10, "bold"))

        # 实验台与桌面分层
        self.scene.create_rectangle(0, 26, cw, ch, fill="#d5d6d8", outline="", width=0)
        self.scene.create_rectangle(0, int(ch * 0.43), cw, ch - 20, fill="#bebfc2", outline="", width=0)
        self.scene.create_rectangle(0, 38, cw, ch - 20, fill="#d0d2d5", outline="#a0a4aa", width=1)


        # 器件布局（支持拖动）
        for dev_id, d in self.devices.items():
            self._add_device(d["x"], d["y"], d["w"], d["h"], d["title"], d["fill"])
            self.scene.create_rectangle(d["x"], d["y"], d["x"] + d["w"], d["y"] + d["h"],
                                        outline="", fill="", tags=("dev_hit", "dev_" + dev_id))

        # 光源与太阳能板细节
        lamp = self.devices["lamp"]
        panel = self.devices["panel"]
        self.scene.create_oval(lamp["x"] + 10, lamp["y"] + 10, lamp["x"] + 38, lamp["y"] + 38,
                               fill="#ffd36a", outline="#ffe8a5", width=1)
        self.scene.create_line(lamp["x"] + 42, lamp["y"] + 16, panel["x"] - 14, panel["y"] + 12, fill="#ffcf66", width=1)
        self.scene.create_line(lamp["x"] + 42, lamp["y"] + 24, panel["x"] - 12, panel["y"] + 26, fill="#ffcf66", width=1)
        self.scene.create_line(lamp["x"] + 42, lamp["y"] + 32, panel["x"] - 14, panel["y"] + 40, fill="#ffcf66", width=1)
        self.scene.create_rectangle(panel["x"] + 18, panel["y"] + 22, panel["x"] + 86, panel["y"] + 56,
                                    fill="#204a68", outline="#4e87af", width=1)
        for gx in (panel["x"] + 34, panel["x"] + 50, panel["x"] + 66):
            self.scene.create_line(gx, panel["y"] + 23, gx, panel["y"] + 55, fill="#6fb4d8", width=1)

        # 仪表/电阻箱细节
        amm = self.devices["amm"]
        res = self.devices["res"]
        vol = self.devices["vol"]
        # 用资源图替换电流表/电压表/电阻箱细节（优先）
        # 图标区固定为设备内部“正方形”区域并等比居中，避免任何方向拉伸
        def _draw_square_icon(dev_id, dev, key, fallback_draw_fn):
            icon_size = max(20, min(int(dev["w"] - 10), int(dev["h"] - 20)))
            icon_x = dev["x"] + (dev["w"] - icon_size) / 2
            icon_y = dev["y"] + 16 + max(0, (dev["h"] - 16 - icon_size) / 2)
            self.device_icon_bounds[dev_id] = (icon_x, icon_y, icon_size)
            icon = self._get_device_asset(key, icon_size, icon_size)
            if icon is not None:
                self.scene.create_image(icon_x, icon_y, anchor="nw", image=icon, tags=("device",))
            else:
                fallback_draw_fn()

        def _draw_amm_fallback():
            self.scene.create_oval(amm["x"] + 14, amm["y"] + 10, amm["x"] + 58, amm["y"] + 44,
                                   fill="#0e1e18", outline="#6dd39b", width=1)
            self.scene.create_text(amm["x"] + 36, amm["y"] + 28, text="A", fill="#8df0bf", font=("Consolas", 9, "bold"))

        def _draw_res_fallback():
            self.scene.create_rectangle(res["x"] + 14, res["y"] + 14, res["x"] + 78, res["y"] + 34,
                                        fill="#0f131d", outline="#ff9f5f", width=1)
            self.scene.create_text(res["x"] + 46, res["y"] + 24, text="R BOX", fill="#ffb27f", font=("Consolas", 8, "bold"))

        def _draw_vol_fallback():
            self.scene.create_oval(vol["x"] + 18, vol["y"] + 12, vol["x"] + 60, vol["y"] + 46,
                                   fill="#1b1836", outline="#9c8cff", width=1)
            self.scene.create_text(vol["x"] + 39, vol["y"] + 29, text="V", fill="#b9abff", font=("Consolas", 9, "bold"))

        _draw_square_icon("amm", amm, "amm", _draw_amm_fallback)
        _draw_square_icon("res", res, "res", _draw_res_fallback)
        _draw_square_icon("vol", vol, "vol", _draw_vol_fallback)

        # 太阳能板端子仍保持设备边缘；仪表与电阻箱端子贴合图标端子位置
        for dev_id, m in self.device_term_offsets.items():
            d = self.devices[dev_id]
            for term_id, (ox, oy) in m.items():
                sign = "+" if term_id.endswith("_p") else "-"
                self._add_terminal(term_id, d["x"] + ox, d["y"] + oy, sign)

        icon_terms = {
            "amm": {"amm_p": (0.22, 0.80), "amm_n": (0.78, 0.80)},
            "vol": {"vol_p": (0.22, 0.80), "vol_n": (0.78, 0.80)},
            # 电阻箱端子对齐到底部三个小圆点中的左右两点
            "res": {"res_p": (0.26, 0.80), "res_n": (0.74, 0.80)},
        }
        for dev_id, tmap in icon_terms.items():
            if dev_id not in self.device_icon_bounds:
                continue
            ix, iy, isz = self.device_icon_bounds[dev_id]
            for term_id, (rx, ry) in tmap.items():
                sign = "+" if term_id.endswith("_p") else "-"
                self._add_terminal(term_id, ix + isz * rx, iy + isz * ry, sign)

        self._redraw_wires()
        self._update_wire_status()

    def _hit_terminal(self, x, y):
        hit = self.scene.find_overlapping(x - 3, y - 3, x + 3, y + 3)
        term = None
        for item in hit:
            tags = self.scene.gettags(item)
            for tag in tags:
                if tag.startswith("term_"):
                    term = tag.replace("term_", "")
                    break
            if term:
                break
        return term

    def _on_scene_press(self, event):
        term = self._hit_terminal(event.x, event.y)
        if term:
            self.drag_start_term = term
            x0, y0 = self.term_pos[term]
            self.drag_line_id = self.scene.create_line(
                x0, y0, event.x, event.y, fill="#ffb26b", width=3,
                capstyle=tk.ROUND, smooth=True, splinesteps=20, tags="wire_preview"
            )
            return
        dev_id = self._hit_device_by_bbox(event.x, event.y)
        if dev_id in ("lamp", "panel"):
            return
        if self.is_distance_experiment and dev_id == "res":
            return
        if dev_id is not None:
            self.drag_device_id = dev_id
            self.drag_device_last_xy = (event.x, event.y)
            return

    def _on_scene_drag(self, event):
        if self.drag_device_id is not None and self.drag_device_last_xy is not None:
            lx, ly = self.drag_device_last_xy
            dx, dy = event.x - lx, event.y - ly
            dev_id = self.drag_device_id
            d = self.devices[dev_id]
            d["x"] += dx
            d["y"] += dy
            self.drag_device_last_xy = (event.x, event.y)
            self._draw_scene()
            # _draw_scene 会重置拖拽状态，这里恢复以保证连续拖动
            self.drag_device_id = dev_id
            self.drag_device_last_xy = (event.x, event.y)
            return
        if self.drag_line_id is None or self.drag_start_term is None:
            return
        x0, y0 = self.term_pos[self.drag_start_term]
        self.scene.coords(self.drag_line_id, x0, y0, event.x, event.y)

    def _on_scene_release(self, event):
        if self.drag_device_id is not None:
            self.drag_device_id = None
            self.drag_device_last_xy = None
            return
        if self.drag_line_id is not None:
            self.scene.delete(self.drag_line_id)
        start = self.drag_start_term
        end = self._hit_terminal(event.x, event.y)
        self.drag_start_term = None
        self.drag_line_id = None
        if not start or not end or start == end:
            return

        a, b = start, end
        pair = tuple(sorted((a, b)))

        if pair in self.wires:
            self.wires.remove(pair)
        else:
            # 简化规则：每个端子最多连接 2 根导线
            if self._terminal_degree(a) >= 2 or self._terminal_degree(b) >= 2:
                self._show_toast("端子连接数已达上限")
                return
            self.wires.add(pair)
        self._redraw_wires()
        self._update_wire_status()

    def _hit_device(self, x, y):
        hit = self.scene.find_overlapping(x, y, x, y)
        for item in hit:
            for tag in self.scene.gettags(item):
                if tag.startswith("dev_") and tag != "dev_hit":
                    dev_id = tag.replace("dev_", "")
                    if dev_id in self.devices:
                        return dev_id
        return self._hit_device_by_bbox(x, y)

    def _hit_device_by_bbox(self, x, y):
        for dev_id, d in self.devices.items():
            if d["x"] <= x <= d["x"] + d["w"] and d["y"] <= y <= d["y"] + d["h"]:
                return dev_id
        return None

    def _on_scene_double_click(self, event):
        dev_id = self._hit_device(event.x, event.y)
        if dev_id in ("res", "panel") and not self.connection_ok:
            self._show_toast("接线未正确，不能调整器材参数")
            return
        if dev_id == "res":
            if self.is_distance_experiment and self.data_points:
                self._show_toast("已有记录数据，请先删除全部记录后再修改电阻")
                return
            self._open_resistance_dialog()
        elif dev_id == "panel" and self.is_distance_experiment:
            if not self.distance_resistance_confirmed:
                self._show_toast("请先双击电阻箱设置阻值，再调节距离")
                return
            self._open_distance_dialog()

    def _set_distance_value(self, dist_cm, auto_record=False):
        self.distance_cm = self._normalize_distance_cm(dist_cm)
        self.light_intensity = self._lookup_distance_light_intensity(self.distance_cm)
        if self.distance_preview_canvas is not None and self.distance_preview_canvas.winfo_exists():
            self._update_distance_preview(self.distance_cm)
        # 在主实验台上同步太阳能板位置，形成“移动板子改距离”的视觉反馈
        lamp_center_x = self.devices["lamp"]["x"] + self.devices["lamp"]["w"] / 2.0
        min_panel_x = 220.0
        max_panel_x = 700.0
        x_new = min_panel_x + (self.distance_cm - 5.0) / 95.0 * (max_panel_x - min_panel_x)
        w_panel = self.devices["panel"]["w"]
        self.devices["panel"]["x"] = int(max(min_panel_x, min(max_panel_x, x_new - w_panel / 2.0)))
        self._draw_scene()
        if auto_record:
            self._try_auto_record_on_distance_change()

    def _normalize_distance_cm(self, dist_cm):
        d = float(dist_cm)
        d = round(d / 5.0) * 5.0
        return max(5.0, min(100.0, d))

    def _open_distance_dialog(self):
        if self.distance_win is not None and self.distance_win.winfo_exists():
            self.distance_win.lift()
            self.distance_win.focus_set()
            return
        dlg = tk.Toplevel(self.frame.winfo_toplevel())
        dlg.title("调节光源-太阳能板距离")
        dlg.configure(bg=BG)
        dlg.geometry("760x250")
        dlg.resizable(False, False)
        self.distance_win = dlg

        tk.Label(dlg, text="距离刻度台（滑动太阳能板）", bg=ACCENT, fg="#fff",
                 font=("Microsoft YaHei", 12, "bold")).pack(fill=tk.X, ipady=6)

        body = tk.Frame(dlg, bg=BG)
        body.pack(fill=tk.BOTH, expand=True, padx=14, pady=10)
        self.distance_preview_canvas = tk.Canvas(body, width=720, height=110, bg="#f6f9ff",
                                                 highlightthickness=1, highlightbackground="#6f8fb4")
        self.distance_preview_canvas.pack(fill=tk.X)
        self._draw_distance_preview_base()
        self._update_distance_preview(self.distance_cm)

        lbl = tk.Frame(body, bg=BG)
        lbl.pack(fill=tk.X, pady=(8, 2))
        tk.Label(lbl, text="距离 d (cm):", bg=BG, fg=FG,
                 font=("Microsoft YaHei", 10, "bold")).pack(side=tk.LEFT)
        self.distance_value_lbl = tk.Label(lbl, text="", bg=BG, fg="#255caa",
                                           font=("Consolas", 12, "bold"))
        self.distance_value_lbl.pack(side=tk.LEFT, padx=8)
        self.distance_vi_lbl = tk.Label(lbl, text="", bg=BG, fg="#0f5132",
                                        font=("Consolas", 12, "bold"))
        self.distance_vi_lbl.pack(side=tk.RIGHT)

        self.distance_scale = tk.Scale(body, from_=5, to=100, orient=tk.HORIZONTAL,
                                       resolution=5, showvalue=False, length=700,
                                       bg=BG, fg=FG, highlightthickness=0,
                                       troughcolor="#bfd4ee",
                                       command=self._on_distance_scale_change)
        self.distance_scale.set(int(self._normalize_distance_cm(self.distance_cm)))
        self.distance_scale.pack(fill=tk.X, pady=(2, 0))
        self._update_distance_reading_preview(self.distance_cm)

        if self.is_distance_experiment:
            ttk.Button(body, text="【实验二】记录当前数据", style="Primary.TButton",
                       command=self._record_point).pack(fill=tk.X, pady=(10, 0), ipady=1)

        dlg.protocol("WM_DELETE_WINDOW", self._close_distance_dialog)

    def _close_distance_dialog(self):
        if self.distance_win is not None and self.distance_win.winfo_exists():
            self.distance_win.destroy()
        self.distance_win = None
        self.distance_scale = None
        self.distance_preview_canvas = None
        self.distance_indicator_id = None

    def _draw_distance_preview_base(self):
        c = self.distance_preview_canvas
        if c is None or not c.winfo_exists():
            return
        c.delete("all")
        c.create_text(32, 20, text="光源", fill="#5c4200", font=("Microsoft YaHei", 10, "bold"), anchor="w")
        c.create_oval(24, 28, 56, 60, fill="#ffd56f", outline="#c48a1f", width=1)
        x0 = 80
        x1 = 680
        y = 72
        c.create_line(x0, y, x1, y, fill="#3b4f6b", width=2)
        for d in range(5, 101, 5):
            t = (d - 5) / 95.0
            x = x0 + t * (x1 - x0)
            tick_h = 16 if d % 10 == 0 else 9
            c.create_line(x, y, x, y - tick_h, fill="#3b4f6b", width=1)
            if d % 10 == 0:
                c.create_text(x, y + 12, text=str(d), fill="#3b4f6b", font=("Consolas", 8))
        c.create_text(x1 + 8, y + 12, text="cm", fill="#3b4f6b", font=("Consolas", 8), anchor="w")
        self.distance_indicator_id = c.create_rectangle(0, 32, 0, 62, fill="#2a6fbb", outline="#1d4f86", width=1)

    def _update_distance_preview(self, dist_cm):
        c = self.distance_preview_canvas
        if c is None or not c.winfo_exists() or self.distance_indicator_id is None:
            return
        x0 = 80
        x1 = 680
        t = (dist_cm - 5.0) / 95.0
        x = x0 + t * (x1 - x0)
        c.coords(self.distance_indicator_id, x - 14, 32, x + 14, 62)
        c.itemconfigure(self.distance_indicator_id)
        c.delete("panel_lbl")
        c.create_text(x, 24, text="太阳能板", fill="#1d4f86", font=("Microsoft YaHei", 9, "bold"), tags="panel_lbl")

    def _update_distance_reading_preview(self, dist_cm):
        v_val, i_val = self._measure_ui_from_distance_data(dist_cm)
        self.distance_value_lbl.config(text="d = {:.0f} cm  |  E = {:.1f} W/m²".format(dist_cm, self.light_intensity))
        self.distance_vi_lbl.config(text="U = {:.3f} V, I = {:.3f} mA".format(v_val, i_val))

    def _on_distance_scale_change(self, val):
        prev_d = self._normalize_distance_cm(self.distance_cm)
        target_d = self._normalize_distance_cm(val)
        if target_d > prev_d:
            d = prev_d + 5.0
        elif target_d < prev_d:
            d = prev_d - 5.0
        else:
            d = prev_d
        d = self._normalize_distance_cm(d)
        self._set_distance_value(d, auto_record=False)
        if self.distance_scale is not None and self.distance_scale.winfo_exists():
            if abs(float(self.distance_scale.get()) - d) > 1e-9:
                self.distance_scale.set(int(d))
            self._update_distance_reading_preview(self.distance_cm)

    def _set_resistance_value(self, val):
        self.r_value = round(max(0.0, float(val)), 1)
        if self.r_value == int(self.r_value):
            txt = "R = {} Ω（双击实验台电阻箱修改）".format(int(self.r_value))
        else:
            txt = "R = {:.1f} Ω（双击实验台电阻箱修改）".format(self.r_value)
        if hasattr(self, "r_readonly_label") and self.r_readonly_label.winfo_exists():
            self.r_readonly_label.config(text=txt)
        self._try_auto_record_on_r_change()

    def _open_resistance_dialog(self):
        dlg = tk.Toplevel(self.frame.winfo_toplevel())
        dlg.title("调整负载电阻")
        dlg.configure(bg=BG)
        dlg.geometry("320x180")
        dlg.resizable(False, False)
        dlg.grab_set()

        tk.Label(dlg, text="电阻箱阻值设定", bg=ACCENT, fg="#fff",
                 font=("Microsoft YaHei", 12, "bold")).pack(fill=tk.X, ipady=6)
        tk.Label(dlg, text="输入阻值 R (Ω)", bg=BG, fg=FG,
                 font=("Microsoft YaHei", 10)).pack(pady=(14, 6))
        v = tk.StringVar(value=str(self.r_value))
        e = tk.Entry(dlg, textvariable=v, font=("Consolas", 12), width=14,
                     bg="#1a1a1a", fg="#00ff88", insertbackground="#00ff88",
                     relief=tk.SUNKEN, bd=2)
        e.pack()
        e.focus_set()
        e.selection_range(0, tk.END)

        def _submit():
            try:
                val = float(v.get().strip())
                if val < 0:
                    raise ValueError
            except ValueError:
                self._show_toast("请输入有效电阻值（>=0）")
                return
            self._set_resistance_value(val)
            if self.is_distance_experiment:
                self.distance_resistance_confirmed = True
            dlg.destroy()

        btns = tk.Frame(dlg, bg=BG)
        btns.pack(fill=tk.X, padx=24, pady=16)
        ttk.Button(btns, text="确认", style="Primary.TButton",
                   command=_submit).pack(side=tk.LEFT, expand=True, padx=4)
        ttk.Button(btns, text="取消", style="Assist.TButton",
                   command=dlg.destroy).pack(side=tk.LEFT, expand=True, padx=4)
        dlg.bind("<Return>", lambda _e: _submit())
        dlg.bind("<Escape>", lambda _e: dlg.destroy())

    def _terminal_degree(self, term_id):
        d = 0
        for x, y in self.wires:
            if x == term_id or y == term_id:
                d += 1
        return d

    def _point_device_id(self, term_id):
        if "_" not in term_id:
            return None
        return term_id.split("_", 1)[0]

    def _line_hits_rect(self, x1, y1, x2, y2, rect):
        rx1, ry1, rx2, ry2 = rect
        # 仅处理横/竖线段（本函数用于折线路径检查）
        if abs(y1 - y2) < 1e-6:
            y = y1
            if y <= ry1 or y >= ry2:
                return False
            a, b = sorted((x1, x2))
            return not (b <= rx1 or a >= rx2)
        if abs(x1 - x2) < 1e-6:
            x = x1
            if x <= rx1 or x >= rx2:
                return False
            a, b = sorted((y1, y2))
            return not (b <= ry1 or a >= ry2)
        return False

    def _polyline_hits_rects(self, pts, rects):
        for i in range(len(pts) - 1):
            x1, y1 = pts[i]
            x2, y2 = pts[i + 1]
            for rect in rects:
                if self._line_hits_rect(x1, y1, x2, y2, rect):
                    return True
        return False

    def _build_wire_polyline(self, a, b, x1, y1, x2, y2):
        # 先尝试简单 L 形，再尝试上下左右绕行，尽量避开仪器
        dev_a = self._point_device_id(a)
        dev_b = self._point_device_id(b)
        margin = 10
        rects = []
        for dev_id, d in self.devices.items():
            if dev_id in (dev_a, dev_b, "lamp"):
                continue
            rects.append((d["x"] - margin, d["y"] - margin, d["x"] + d["w"] + margin, d["y"] + d["h"] + margin))

        p_hv = [(x1, y1), (x2, y1), (x2, y2)]
        p_vh = [(x1, y1), (x1, y2), (x2, y2)]
        if not self._polyline_hits_rects(p_hv, rects):
            return p_hv
        if not self._polyline_hits_rects(p_vh, rects):
            return p_vh

        cw, ch = self.scene_size
        # 多给几条候选走廊：上、下、左、右
        y_top = 42
        y_bot = ch - 34
        x_left = 24
        x_right = cw - 24
        candidates = [
            [(x1, y1), (x1, y_top), (x2, y_top), (x2, y2)],
            [(x1, y1), (x1, y_bot), (x2, y_bot), (x2, y2)],
            [(x1, y1), (x_left, y1), (x_left, y2), (x2, y2)],
            [(x1, y1), (x_right, y1), (x_right, y2), (x2, y2)],
        ]
        for pts in candidates:
            if not self._polyline_hits_rects(pts, rects):
                return pts
        # 兜底：返回较自然的 L 形
        return p_hv if abs(x2 - x1) >= abs(y2 - y1) else p_vh

    def _redraw_wires(self):
        if self.scene is None or not self.scene.winfo_exists():
            return
        self.scene.delete("wire")
        for a, b in self.wires:
            x1, y1 = self.term_pos[a]
            x2, y2 = self.term_pos[b]
            self.scene.create_line(
                x1, y1, x2, y2,
                fill="#6a2a14", width=5.2,
                capstyle=tk.ROUND, joinstyle=tk.ROUND, tags="wire"
            )
            self.scene.create_line(
                x1, y1, x2, y2,
                fill="#ff7b3d", width=3.2,
                capstyle=tk.ROUND, joinstyle=tk.ROUND, tags="wire"
            )
            # 端子插头点
            self.scene.create_oval(x1 - 2.8, y1 - 2.8, x1 + 2.8, y1 + 2.8,
                                   fill="#ffd38a", outline="#70451d", width=1, tags="wire")
            self.scene.create_oval(x2 - 2.8, y2 - 2.8, x2 + 2.8, y2 + 2.8,
                                   fill="#ffd38a", outline="#70451d", width=1, tags="wire")

    def _update_wire_status(self):
        if self.is_distance_experiment:
            if self.exp2_measure_mode == "isc":
                required = {
                    tuple(sorted(("panel_p", "amm_p"))),
                    tuple(sorted(("amm_n", "panel_n"))),
                }
                forbidden = {
                    tuple(sorted(("amm_n", "res_p"))),
                    tuple(sorted(("res_n", "panel_n"))),
                    tuple(sorted(("vol_p", "res_p"))),
                    tuple(sorted(("vol_n", "res_n"))),
                    tuple(sorted(("vol_p", "panel_p"))),
                    tuple(sorted(("vol_n", "panel_n"))),
                }
                ok = required.issubset(self.wires) and not any(w in self.wires for w in forbidden)
            else:
                required = {
                    tuple(sorted(("vol_p", "panel_p"))),
                    tuple(sorted(("vol_n", "panel_n"))),
                }
                forbidden = {
                    tuple(sorted(("panel_p", "amm_p"))),
                    tuple(sorted(("amm_n", "panel_n"))),
                    tuple(sorted(("amm_n", "res_p"))),
                    tuple(sorted(("res_n", "panel_n"))),
                    tuple(sorted(("vol_p", "res_p"))),
                    tuple(sorted(("vol_n", "res_n"))),
                }
                ok = required.issubset(self.wires) and not any(w in self.wires for w in forbidden)
        else:
            required = {
                tuple(sorted(("panel_p", "amm_p"))),
                tuple(sorted(("amm_n", "res_p"))),
                tuple(sorted(("res_n", "panel_n"))),
                tuple(sorted(("vol_p", "res_p"))),
                tuple(sorted(("vol_n", "res_n"))),
            }
            ok = required.issubset(self.wires)
        self.connection_ok = ok
        if self.wire_status is None or not self.wire_status.winfo_exists():
            self._update_circuit_diagram()
            return
        if ok:
            if self.is_distance_experiment:
                mode_text = "短路电流 Isc" if self.exp2_measure_mode == "isc" else "开路电压 Voc"
                self.wire_status.config(text="接线状态: 正确（{}）".format(mode_text), fg="#0d7c3b")
            else:
                self.wire_status.config(text="接线状态: 正确，可采集", fg="#0d7c3b")
                self._try_auto_record_on_r_change()
        else:
            if self.is_distance_experiment:
                self.wire_status.config(text="接线状态: 未连线", fg="#9a4f00")
            else:
                self.wire_status.config(text="接线状态: 未完成", fg="#9a4f00")
        self._update_circuit_diagram()

    def _auto_wire_for_test(self):
        """测试阶段：一键完成实验一标准接线。"""
        if self.is_distance_experiment:
            if self.exp2_measure_mode == "isc":
                self.wires = {
                    tuple(sorted(("panel_p", "amm_p"))),
                    tuple(sorted(("amm_n", "panel_n"))),
                }
            else:
                self.wires = {
                    tuple(sorted(("vol_p", "panel_p"))),
                    tuple(sorted(("vol_n", "panel_n"))),
                }
        else:
            self.wires = {
                tuple(sorted(("panel_p", "amm_p"))),
                tuple(sorted(("amm_n", "res_p"))),
                tuple(sorted(("res_n", "panel_n"))),
                tuple(sorted(("vol_p", "res_p"))),
                tuple(sorted(("vol_n", "res_n"))),
            }
        self._redraw_wires()
        self._update_wire_status()
        self._show_toast("已自动完成测试接线")

    def _auto_wire_exp2_isc(self):
        self._set_exp2_measure_mode("isc")
        self._auto_wire_for_test()

    def _auto_wire_exp2_voc(self):
        self._set_exp2_measure_mode("voc")
        self._auto_wire_for_test()

    def _set_exp2_measure_mode(self, mode):
        if not self.is_distance_experiment:
            return
        if mode not in ("isc", "voc"):
            return
        self.exp2_measure_mode = mode
        if mode == "isc":
            self._set_resistance_value(0.0)
        self.connection_ok = False
        self._update_wire_status()
        self._show_toast("已切换到 {} 接线模式".format("短路电流 Isc" if mode == "isc" else "开路电压 Voc"))

    def _layout_devices_for_large_scene(self):
        self.devices["lamp"]["x"], self.devices["lamp"]["y"] = 70, 70
        self.devices["panel"]["x"], self.devices["panel"]["y"] = 380, 60
        self.devices["amm"]["x"], self.devices["amm"]["y"] = 70, 240
        self.devices["res"]["x"], self.devices["res"]["y"] = 420, 250
        self.devices["vol"]["x"], self.devices["vol"]["y"] = 220, 420

    def _update_circuit_diagram(self):
        if self.circuit_canvas is None or not self.circuit_canvas.winfo_exists():
            return
        c = self.circuit_canvas
        c.delete("all")
        c.create_rectangle(0, 0, 320, 250, fill="#f7f4e8", outline="#20242d", width=1)
        c.create_text(160, 16, text="当前接线图", fill="#111",
                      font=("Microsoft YaHei", 10, "bold"))

        p = {
            "panel_p": (96, 88), "panel_n": (96, 128),
            "amm_p": (132, 88), "amm_n": (176, 88),
            "res_p": (218, 88), "res_n": (286, 88),
            "vol_p": (218, 174), "vol_n": (286, 174),
        }

        # 先画实际导线：连了哪两个端子，就画哪两个端子之间的直线。
        for a, b in sorted(self.wires):
            if a not in p or b not in p:
                continue
            x1, y1 = p[a]
            x2, y2 = p[b]
            c.create_line(x1, y1, x2, y2, fill="#111", width=2.4, capstyle=tk.ROUND)

        # 电池板
        c.create_rectangle(28, 62, 96, 154, outline="#111", width=2, fill="#fffdf2")
        for gx in (45, 62, 79):
            c.create_line(gx, 68, gx, 148, fill="#444", width=1)
        for gy in (92, 124):
            c.create_line(34, gy, 90, gy, fill="#444", width=1)
        c.create_text(62, 174, text="电池板", fill="#111", font=("Microsoft YaHei", 9, "bold"))

        # A 圈
        c.create_oval(132, 66, 176, 110, outline="#111", width=2, fill="#fffdf2")
        c.create_text(154, 88, text="A", fill="#111", font=("Consolas", 15, "bold"))

        # 电阻箱
        c.create_rectangle(218, 62, 286, 114, outline="#111", width=2, fill="#fffdf2")
        c.create_text(252, 88, text="电阻箱", fill="#111", font=("Microsoft YaHei", 9, "bold"))

        # V 圈
        c.create_oval(230, 152, 274, 196, outline="#111", width=2, fill="#fffdf2")
        c.create_text(252, 174, text="V", fill="#111", font=("Consolas", 15, "bold"))

        for tid, (x, y) in p.items():
            connected = any(tid in pair for pair in self.wires)
            c.create_oval(x - 3, y - 3, x + 3, y + 3,
                          fill="#111" if connected else "#fffdf2",
                          outline="#111",
                          width=1)
        c.tkraise()

    def _update_core_principle(self, U_val, I_mA):
        """核心链路：Pin = E*A，Pout = U*I，eta = Pout/Pin。"""
        pin = self.light_intensity * self.panel_area_m2
        pout = U_val * I_mA / 1000.0
        eta = (pout / pin * 100.0) if pin > 0 else 0.0
        if hasattr(self, "core_pin_lbl") and self.core_pin_lbl is not None and self.core_pin_lbl.winfo_exists():
            self.core_pin_lbl.config(text="入射光功率 Pin = {:.4f} W".format(pin))
        if hasattr(self, "core_pout_lbl") and self.core_pout_lbl is not None and self.core_pout_lbl.winfo_exists():
            self.core_pout_lbl.config(text="输出电功率 Pout = {:.4f} W".format(pout))
        if hasattr(self, "core_eta_lbl") and self.core_eta_lbl is not None and self.core_eta_lbl.winfo_exists():
            self.core_eta_lbl.config(text="转换效率 η = {:.2f} %".format(eta))

    def _build_resistance_box(self, parent):
        tk.Label(parent, text="━━━ 负载电阻 ━━━", bg=PANEL_BG, fg=ACCENT,
                 font=("Microsoft YaHei", 10, "bold")).pack(anchor="w", padx=10, pady=(0, 2))
        box = tk.Frame(parent, bg="#1e1e1e", relief=tk.RAISED, bd=1)
        box.pack(fill=tk.X, padx=10, pady=4)

        tk.Label(box, text="电阻箱 (Ω)", bg="#1e1e1e", fg="#ff6b35",
                 font=("Microsoft YaHei", 9, "bold")).pack(pady=(4, 2))

        # 旋钮区
        knobs = tk.Frame(box, bg="#1e1e1e")
        knobs.pack(padx=2)
        self.r_digits = []
        self.r_multipliers = [10000, 1000, 100, 10, 1, 0.1]
        labels = ["×10k", "×1k", "×100", "×10", "×1", "×0.1"]
        for i, lbl in enumerate(labels):
            col = tk.Frame(knobs, bg="#2b2b2b")
            col.pack(side=tk.LEFT, padx=1)
            tk.Label(col, text=lbl, bg="#2b2b2b", fg="#aaa",
                     font=("Consolas", 7)).pack()
            ttk.Button(col, text="▲", style="Assist.TButton",
                       command=lambda idx=i: self._r_digit_inc(idx)).pack(pady=1)
            digit_lbl = tk.Label(col, text="0", bg="#1a1a1a", fg="#ff6b35",
                                 width=2, font=("Consolas", 13, "bold"))
            digit_lbl.pack(pady=1)
            ttk.Button(col, text="▼", style="Assist.TButton",
                       command=lambda idx=i: self._r_digit_dec(idx)).pack(pady=1)
            self.r_digits.append(digit_lbl)

        # 总电阻显示
        self.r_total_label = tk.Label(box, text="R = 0.0 Ω", bg="#1e1e1e",
                                      fg="#00ff88", font=("Consolas", 12, "bold"))
        self.r_total_label.pack(pady=(4, 4))
        self.r_value = 0.0

        # 手动输入电阻值
        manual_f = tk.Frame(box, bg="#1e1e1e")
        manual_f.pack(fill=tk.X, padx=8, pady=(0, 6))
        tk.Label(manual_f, text="直接输入:", bg="#1e1e1e", fg="#aaa",
                 font=("Microsoft YaHei", 8)).pack(side=tk.LEFT)
        self.r_manual_var = tk.StringVar(value="")
        r_manual_entry = tk.Entry(manual_f, textvariable=self.r_manual_var,
                                  font=("Consolas", 11), bg="#1a1a1a", fg="#00ff88",
                                  insertbackground="#00ff88", width=8,
                                  relief=tk.SUNKEN, bd=1)
        r_manual_entry.pack(side=tk.LEFT, padx=4)
        r_manual_entry.bind("<Return>", self._on_r_manual_enter)
        tk.Label(manual_f, text="Ω", bg="#1e1e1e", fg="#aaa",
                 font=("Consolas", 9)).pack(side=tk.LEFT)
        ttk.Button(manual_f, text="设定", style="Assist.TButton",
                   command=lambda: self._on_r_manual_enter(None)).pack(side=tk.LEFT, padx=2)

    # ── 电阻箱旋钮操作 ──

    def _r_digit_inc(self, idx):
        lbl = self.r_digits[idx]
        v = (int(lbl.cget("text")) + 1) % 10
        lbl.config(text=str(v))
        self._r_update_total()

    def _r_digit_dec(self, idx):
        lbl = self.r_digits[idx]
        v = (int(lbl.cget("text")) - 1) % 10
        lbl.config(text=str(v))
        self._r_update_total()

    def _r_update_total(self):
        total = 0.0
        for i, mult in enumerate(self.r_multipliers):
            total += int(self.r_digits[i].cget("text")) * mult
        total = round(total, 1)
        self.r_value = total
        if total == int(total):
            self.r_total_label.config(text="R = {} Ω".format(int(total)))
        else:
            self.r_total_label.config(text="R = {:.1f} Ω".format(total))
        self.r_manual_var.set("")
        self._try_auto_record_on_r_change()

    def _on_r_manual_enter(self, event):
        """手动输入电阻值"""
        try:
            val = float(self.r_manual_var.get().strip())
            if val < 0:
                raise ValueError
        except ValueError:
            self._show_toast("请输入有效的电阻值（≥0）")
            return
        self.r_value = val
        if val == int(val):
            self.r_total_label.config(text="R = {} Ω".format(int(val)))
        else:
            self.r_total_label.config(text="R = {:.1f} Ω".format(val))
        # 同步电阻箱位数显示（按位拆分，避免出现 99 Ω 显示成 109 Ω 的问题）
        remaining = val
        for i, mult in enumerate(self.r_multipliers):
            if mult >= 1:
                digit = int(remaining // mult)
                remaining -= digit * mult
            else:
                digit = int(round(remaining / mult))
            digit = max(0, min(9, digit))
            self.r_digits[i].config(text=str(digit))
        self._try_auto_record_on_r_change()

    # ── 电池类型切换 ──

    def _on_type_change(self):
        t = self.cell_type_var.get()
        self.params = {k: v for k, v in DEFAULT_PARAMS[t].items()}

    # ── 右侧：两张图 + 数据表 ──

    def _build_right(self, parent):
        btns = tk.Frame(parent, bg=PANEL_BG)
        btns.pack(fill=tk.X, padx=8, pady=(8, 6))
        ttk.Button(btns, text="打开特性曲线窗口", style="Primary.TButton",
                   command=self._open_curve_window).pack(fill=tk.X, ipady=2)
        if not self.is_distance_experiment:
            ttk.Button(btns, text="数据分析", style="Info.TButton",
                       command=self._show_analysis).pack(fill=tk.X, pady=(6, 0), ipady=2)

        btn_row = tk.Frame(parent, bg=PANEL_BG)
        btn_row.pack(fill=tk.X, padx=8, pady=(6, 8))
        ttk.Button(btn_row, text="打开实验数据表", style="Info.TButton",
                   command=self._open_data_table_window).pack(fill=tk.X, ipady=2)
        self._build_autowire_buttons(parent, pady=(2, 8))

    def _open_data_table_window(self):
        if self.table_win is not None and self.table_win.winfo_exists():
            self.table_win.lift()
            self.table_win.focus_set()
            return
        win = tk.Toplevel(self.frame.winfo_toplevel())
        win.title("{}数据表".format(self.experiment_name))
        win.configure(bg=BG)
        win.geometry("620x360")
        win.resizable(True, True)
        self.table_win = win

        hdr = tk.Frame(win, bg=PANEL_BG)
        hdr.pack(fill=tk.X, padx=8, pady=(8, 4))
        tk.Label(hdr, text="实验数据记录", bg=PANEL_BG, fg=ACCENT,
                 font=("Microsoft YaHei", 11, "bold")).pack(side=tk.LEFT)
        ttk.Button(hdr, text="删除选中行", style="Assist.TButton",
                   command=self._delete_selected).pack(side=tk.RIGHT, padx=4)

        table_frame = tk.Frame(win, bg=PANEL_BG)
        table_frame.pack(fill=tk.BOTH, expand=True, padx=8, pady=(0, 8))
        cols = ("序号", "d (cm)", "I (W/m²)", "VOC (V)", "ISC (mA)") if self.is_distance_experiment else ("序号", "R (Ω)", "U (V)", "I (mA)", "P (mW)")
        self.tree = ttk.Treeview(table_frame, columns=cols, show="headings")
        for c in cols:
            self.tree.heading(c, text=c)
            w = 60 if c == "序号" else 115
            self.tree.column(c, width=w, anchor="center")
        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        self.tree.bind("<Double-1>", self._on_table_double_click)
        self._refresh_table_view()

        def _on_close():
            self.tree = None
            self.table_win = None
            win.destroy()

        win.protocol("WM_DELETE_WINDOW", _on_close)

    def _open_curve_window(self):
        if self.curve_win is not None and self.curve_win.winfo_exists():
            self.curve_win.lift()
            self.curve_win.focus_set()
            return
        win = tk.Toplevel(self.frame.winfo_toplevel())
        win.title("{}特性曲线".format(self.experiment_name))
        win.configure(bg=BG)
        win.geometry("980x460")
        win.resizable(True, True)
        self.curve_win = win

        chart_frame = tk.Frame(win, bg=BG)
        chart_frame.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)
        col1 = tk.Frame(chart_frame, bg=BG)
        col1.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 4))
        col2 = tk.Frame(chart_frame, bg=BG)
        col2.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(4, 0))
        self.fig1 = Figure(figsize=(4.5, 3.2), dpi=100, facecolor=BG)
        self.ax1 = self.fig1.add_subplot(111)
        self._style_ax(self.ax1, title="伏安特性曲线", xlabel="U (V)", ylabel="I (mA)")
        self.canvas1 = FigureCanvasTkAgg(self.fig1, master=col1)
        self.canvas1.get_tk_widget().pack(fill=tk.BOTH, expand=True)
        self.curve_btn1 = ttk.Button(col1, text="展示拟合曲线", style="Assist.TButton",
                                     command=lambda: self._toggle_fit_curve(1))
        self.curve_btn1.pack(fill=tk.X, pady=(6, 0), ipady=3)

        self.fig2 = Figure(figsize=(4.5, 3.2), dpi=100, facecolor=BG)
        self.ax2 = self.fig2.add_subplot(111)
        self._style_ax(self.ax2, title="功率输出曲线（P-V）", xlabel="U (V)", ylabel="P (mW)")
        self.canvas2 = FigureCanvasTkAgg(self.fig2, master=col2)
        self.canvas2.get_tk_widget().pack(fill=tk.BOTH, expand=True)
        self.curve_btn2 = ttk.Button(col2, text="展示拟合曲线", style="Assist.TButton",
                                     command=lambda: self._toggle_fit_curve(2))
        self.curve_btn2.pack(fill=tk.X, pady=(6, 0), ipady=3)
        min_points = 2 if self.is_distance_experiment else 10
        if len(self.data_points) >= min_points:
            self._draw_plot()

        def _on_close():
            self.curve_win = None
            self.fig1 = self.fig2 = None
            self.ax1 = self.ax2 = None
            self.canvas1 = self.canvas2 = None
            self.curve_btn1 = self.curve_btn2 = None
            self.curve_fit_mode_1 = False
            self.curve_fit_mode_2 = False
            win.destroy()

        win.protocol("WM_DELETE_WINDOW", _on_close)

    def _open_curve_and_analysis(self):
        self._open_curve_window()
        if not self.is_distance_experiment:
            self._show_analysis()

    def _toggle_fit_curve(self, which):
        if which == 1:
            self.curve_fit_mode_1 = not self.curve_fit_mode_1
        else:
            self.curve_fit_mode_2 = not self.curve_fit_mode_2
        self._draw_plot(auto_open_window=False)

    def _build_autowire_buttons(self, parent, pady=(6, 6)):
        # 主页不再显示接线/导入快捷按钮，统一放到顶部“工具箱”
        return

    def _refresh_table_view(self):
        if self.tree is None or not self.tree.winfo_exists():
            return
        for item_id in self.tree.get_children():
            self.tree.delete(item_id)
        if self.is_distance_experiment:
            rows = []
            for dp in self.data_points:
                rows.append({
                    "d": float(dp["d"]),
                    "P": float(dp["P"]),
                    "U": float(dp["U"]),
                    "I": float(dp["I"]),
                })
            for d, pp in self.exp2_partial_points.items():
                d = float(d)
                if any(abs(r["d"] - d) < 1e-9 for r in rows):
                    continue
                rows.append({
                    "d": d,
                    "P": float(pp.get("P", self._lookup_distance_light_intensity(d))),
                    "U": pp.get("U"),
                    "I": pp.get("I"),
                })
            rows.sort(key=lambda r: r["d"])
            for idx, r in enumerate(rows, start=1):
                u_txt = "{:.3f}".format(r["U"]) if r["U"] is not None else "--"
                i_txt = "{:.3f}".format(r["I"]) if r["I"] is not None else "--"
                self.tree.insert("", "end", values=(
                    idx, "{:g}".format(r["d"]), "{:.3f}".format(r["P"]), u_txt, i_txt
                ))
        else:
            for idx, dp in enumerate(self.data_points, start=1):
                self.tree.insert("", "end", values=(
                    idx, "{:g}".format(dp["R"]), "{:.3f}".format(dp["U"]),
                    "{:.3f}".format(dp["I"]), "{:.3f}".format(dp["P"])
                ))

    def _style_ax(self, ax, title="", xlabel="", ylabel=""):
        ax.set_facecolor("#0f0f23")
        ax.set_title(title, fontproperties=plt_font, color=FG, fontsize=11, pad=6)
        ax.set_xlabel(xlabel, fontproperties=plt_font, color=FG, fontsize=9)
        ax.set_ylabel(ylabel, fontproperties=plt_font, color=FG, fontsize=9)
        ax.tick_params(colors="#888", labelsize=8)
        for spine in ax.spines.values():
            spine.set_color(GRID_COLOR)
        ax.grid(True, alpha=0.2, color=GRID_COLOR)

    def _best_point_index(self):
        """最佳点规则：P 最大；若 P 在容差内相同，取 R 更小者。"""
        if not self.data_points:
            return None
        eps = 1e-9
        best_i = 0
        best_p = self.data_points[0]["P"]
        first_x = self.data_points[0]["d"] if self.is_distance_experiment else self.data_points[0]["R"]
        best_r = first_x
        for i in range(1, len(self.data_points)):
            p = self.data_points[i]["P"]
            r = self.data_points[i]["d"] if self.is_distance_experiment else self.data_points[i]["R"]
            if p > best_p + eps:
                best_i, best_p, best_r = i, p, r
            elif abs(p - best_p) <= eps and r < best_r:
                best_i, best_r = i, r
        return best_i

    # ── 记录数据（弹窗手动输入 U、I）──

    def _record_point(self):
        self._record_point_impl(show_feedback=True)

    def _record_point_impl(self, show_feedback=False):
        if not self.connection_ok:
            self._show_toast("请先按场景完成正确接线")
            return
        if self.is_distance_experiment and self.exp2_measure_mode == "isc" and abs(self.r_value) > 1e-9:
            self._show_toast("Isc 测量时电阻必须为 0 Ω")
            return
        x_key = "d" if self.is_distance_experiment else "R"
        x_val = self.distance_cm if self.is_distance_experiment else self.r_value
        if not self.is_distance_experiment:
            for dp in self.data_points:
                if abs(dp[x_key] - x_val) < 1e-6:
                    label = "d={:g} cm".format(x_val) if self.is_distance_experiment else "R={:g} Ω".format(x_val)
                    self._show_toast("已存在 {} 的数据，不可重复记录！".format(label))
                    return

        U_val, I_val = self._measure_ui_from_experimental_data(x_val)
        P_val = self.light_intensity if self.is_distance_experiment else (U_val * I_val)

        if self.is_distance_experiment:
            key = float(x_val)
            for dp in self.data_points:
                if abs(dp["d"] - key) < 1e-6:
                    self._show_toast("已存在 d={:g} cm 的完整数据，不可重复记录".format(key))
                    return
            partial = self.exp2_partial_points.get(key, {"d": key, "P": P_val})
            partial["P"] = P_val
            if self.exp2_measure_mode == "isc":
                partial["I"] = I_val
                self.exp2_partial_points[key] = partial
                if self.tree is not None and self.tree.winfo_exists():
                    self._refresh_table_view()
                self._show_toast("已记录 Isc（d={:g} cm），请切换到 Voc 接线后再记录".format(key))
                return
            partial["U"] = U_val
            if "I" not in partial:
                self.exp2_partial_points[key] = partial
                if self.tree is not None and self.tree.winfo_exists():
                    self._refresh_table_view()
                self._show_toast("已记录 Voc（d={:g} cm），请切换到 Isc 接线后再记录".format(key))
                return
            point = {"U": partial["U"], "I": partial["I"], "P": partial["P"], "d": key}
            self.exp2_partial_points.pop(key, None)
        else:
            point = {"U": U_val, "I": I_val, "P": P_val}
            point[x_key] = x_val
        self.data_points.append(point)
        self._update_core_principle(U_val, I_val)

        if self.tree is not None and self.tree.winfo_exists():
            self._refresh_table_view()
        if self.is_distance_experiment:
            self.last_auto_record_distance = x_val
        else:
            self.last_auto_record_r = x_val
        if show_feedback:
            label = "d={:g} cm".format(x_val) if self.is_distance_experiment else "R={:g} Ω".format(x_val)
            self._show_toast("已记录 {}".format(label))

    def _measure_ui_from_experimental_data(self, x_val):
        """实验采集：命中标准点直接返回；否则按报告结论反推标准结果。"""
        if self.is_distance_experiment:
            self._set_distance_value(x_val, auto_record=False)
            v_val, i_val = self._measure_ui_from_distance_data(self.distance_cm)
            return float(v_val), float(i_val)

        return self._infer_standard_iv_from_report(float(x_val))

    def _infer_standard_iv_from_report(self, R):
        """实验一：按报告标准数据 + 结论(Isc/Uoc/FF/R0)反推标准结果。"""
        data = get_experimental_data()["iv_data"]
        Rs = [float(x) for x in data["R"]]
        Us = [float(x) for x in data["U"]]
        Is = [float(x) for x in data["I"]]

        # 命中标准点：直接返回标准结果
        for r0, u0, i0 in zip(Rs, Us, Is):
            if abs(R - r0) < 1e-9:
                return float(u0), float(i0)

        # 非标准点：先按报告表插值
        order = np.argsort(Rs)
        Rs_s = [Rs[i] for i in order]
        Us_s = [Us[i] for i in order]
        Is_s = [Is[i] for i in order]

        if R <= Rs_s[0]:
            u_raw, i_raw = Us_s[0], Is_s[0]
        elif R >= Rs_s[-1]:
            u_raw, i_raw = Us_s[-1], Is_s[-1]
        else:
            u_raw, i_raw = Us_s[-1], Is_s[-1]
            for i in range(len(Rs_s) - 1):
                r1, r2 = Rs_s[i], Rs_s[i + 1]
                if r1 <= R <= r2:
                    t = (R - r1) / (r2 - r1) if r2 != r1 else 0.0
                    u_raw = Us_s[i] + t * (Us_s[i + 1] - Us_s[i])
                    i_raw = Is_s[i] + t * (Is_s[i + 1] - Is_s[i])
                    break

        # 按报告结论做归一：Isc/Uoc/FF/R0
        Isc_std = float(get_experimental_data()["Isc_mA"])
        Uoc_std = float(get_experimental_data()["Uoc_V"])
        FF_std = float(get_experimental_data()["FF"])
        R0_std = float(get_experimental_data()["R0_ohm"])

        Isc_tab = max(Is_s) if Is_s else Isc_std
        Uoc_tab = max(Us_s) if Us_s else Uoc_std
        su = (Uoc_std / Uoc_tab) if Uoc_tab > 1e-12 else 1.0
        si = (Isc_std / Isc_tab) if Isc_tab > 1e-12 else 1.0
        u_adj = float(u_raw) * su
        i_adj = float(i_raw) * si

        # 用 R0 与 FF 对功率峰值做局部修正，避免偏离报告结论
        p_target = FF_std * Isc_std * Uoc_std
        u_r0, i_r0 = self._interp_by_distance(R0_std, Rs_s, Us_s), self._interp_by_distance(R0_std, Rs_s, Is_s)
        p_r0 = (float(u_r0) * su) * (float(i_r0) * si)
        if p_r0 > 1e-12:
            alpha = max(0.6, min(1.6, p_target / p_r0))
            sigma = max(18.0, 0.25 * max(R0_std, 1.0))
            gain = 1.0 + (alpha - 1.0) * np.exp(-((R - R0_std) / sigma) ** 2)
            i_adj *= gain

        return max(0.0, float(u_adj)), max(0.0, float(i_adj))

    def _measure_ui_from_distance_data(self, dist_cm):
        """实验二采集：命中标准点直接返回；否则按报告数据插值反推标准结果。"""
        data = get_experimental_data()["distance_data"]
        ds = [float(x) for x in data["d_cm"]]
        vs = [float(x) for x in data["Voc_V"]]
        cs = [float(x) for x in data["Isc_mA"]]
        for d0, v0, c0 in zip(ds, vs, cs):
            if abs(dist_cm - d0) < 1e-9:
                return float(v0), float(c0)
        voc = self._interp_by_distance(dist_cm, ds, vs)
        isc = self._interp_by_distance(dist_cm, ds, cs)
        return float(voc), float(isc)

    def _lookup_distance_light_intensity(self, dist_cm):
        data = get_experimental_data()["distance_data"]
        ds = [float(x) for x in data["d_cm"]]
        es = [float(x) for x in data["E_wm2"]]
        return float(self._interp_by_distance(dist_cm, ds, es))

    def _interp_by_distance(self, d, xs, ys):
        for x0, y0 in zip(xs, ys):
            if abs(d - x0) < 1e-9:
                return y0
        order = np.argsort(xs)
        xs_s = [xs[i] for i in order]
        ys_s = [ys[i] for i in order]
        # 区间外采用外推（优先 log-log，符合距离类实验的幂律趋势），避免被边界值钳死
        if d <= xs_s[0]:
            x1, x2 = xs_s[0], xs_s[1]
            y1, y2 = ys_s[0], ys_s[1]
            if d <= 0:
                d = 0.1
            if x1 > 0 and x2 > 0 and y1 > 0 and y2 > 0:
                k = (np.log(y2) - np.log(y1)) / (np.log(x2) - np.log(x1))
                b = np.log(y1) - k * np.log(x1)
                return float(np.exp(k * np.log(d) + b))
            t = (d - x1) / (x2 - x1) if x2 != x1 else 0.0
            return ys_s[0] + t * (ys_s[1] - ys_s[0])
        if d >= xs_s[-1]:
            x1, x2 = xs_s[-2], xs_s[-1]
            y1, y2 = ys_s[-2], ys_s[-1]
            if x1 > 0 and x2 > 0 and y1 > 0 and y2 > 0:
                k = (np.log(y2) - np.log(y1)) / (np.log(x2) - np.log(x1))
                b = np.log(y2) - k * np.log(x2)
                return float(np.exp(k * np.log(d) + b))
            t = (d - x2) / (x2 - x1) if x2 != x1 else 0.0
            return ys_s[-1] + t * (ys_s[-1] - ys_s[-2])
        for i in range(len(xs_s) - 1):
            x1, x2 = xs_s[i], xs_s[i + 1]
            if x1 <= d <= x2:
                t = (d - x1) / (x2 - x1) if x2 != x1 else 0.0
                return ys_s[i] + t * (ys_s[i + 1] - ys_s[i])
        return ys_s[-1]

    def _try_auto_record_on_r_change(self):
        """接线正确后，调电阻自动采集并记录（同一 R 仅记录一次）。"""
        if not self.connection_ok:
            return
        if self.is_distance_experiment:
            return
        if self.last_auto_record_r is not None and abs(self.r_value - self.last_auto_record_r) < 1e-9:
            return
        for dp in self.data_points:
            if abs(dp["R"] - self.r_value) < 1e-9:
                return
        self._record_point_impl(show_feedback=False)

    def _try_auto_record_on_distance_change(self):
        # 实验二改为手动点击按钮记录，不再随距离变化自动记录。
        return

    # ── 删除选中行 ──

    def _delete_selected(self):
        if self.tree is None or not self.tree.winfo_exists():
            self._show_toast("请先打开实验数据表")
            return
        sel = self.tree.selection()
        if not sel:
            self._show_toast("请先在表格中选中要删除的行")
            return
        for item_id in sel:
            vals = self.tree.item(item_id, "values")
            x_val = float(vals[1])
            if self.is_distance_experiment:
                self.data_points = [dp for dp in self.data_points if abs(dp["d"] - x_val) > 1e-6]
                self.exp2_partial_points.pop(float(x_val), None)
            else:
                self.data_points = [dp for dp in self.data_points if abs(dp["R"] - x_val) > 1e-6]
            self.tree.delete(item_id)
        self._reindex_table()

    def _reindex_table(self):
        if self.tree is None or not self.tree.winfo_exists():
            return
        for i, item_id in enumerate(self.tree.get_children()):
            vals = list(self.tree.item(item_id, "values"))
            vals[0] = i + 1
            self.tree.item(item_id, values=vals)

    def _on_table_double_click(self, event):
        if self.tree is None or not self.tree.winfo_exists():
            return
        if self.is_distance_experiment:
            self._show_toast("实验二记录请通过接线与记录按钮更新")
            return
        item_id = self.tree.identify_row(event.y)
        col_id = self.tree.identify_column(event.x)
        if not item_id or not col_id:
            return

        # 仅允许编辑 X/U/I 三列（实验二对应 X/VOC/ISC）；序号和派生列不可编辑
        blocked_cols = ("#1", "#3") if self.is_distance_experiment else ("#1", "#5")
        if col_id in blocked_cols:
            self._show_toast("该列不可编辑")
            return

        x_key = "d" if self.is_distance_experiment else "R"
        x_name = "d (cm)" if self.is_distance_experiment else "R (Ω)"
        if self.is_distance_experiment:
            cols_map = {"#2": (x_key, x_name), "#4": ("U", "VOC (V)"), "#5": ("I", "ISC (mA)")}
        else:
            cols_map = {"#2": (x_key, x_name), "#3": ("U", "U (V)"), "#4": ("I", "I (mA)")}
        if col_id not in cols_map:
            return
        field, field_text = cols_map[col_id]

        items = self.tree.get_children()
        row_idx = items.index(item_id)
        old_val = self.data_points[row_idx][field]

        dlg = tk.Toplevel(self.frame.winfo_toplevel())
        dlg.title("修改数据")
        dlg.configure(bg=BG)
        dlg.geometry("320x170")
        dlg.resizable(False, False)
        dlg.grab_set()

        tk.Label(dlg, text="编辑 {}".format(field_text), bg=ACCENT, fg="#fff",
                 font=("Microsoft YaHei", 12, "bold")).pack(fill=tk.X, ipady=6)
        tk.Label(dlg, text="原值: {:.3f}".format(old_val), bg=BG, fg="#aaa",
                 font=("Consolas", 10)).pack(pady=(14, 8))

        v = tk.StringVar(value=str(old_val))
        e = tk.Entry(dlg, textvariable=v, font=("Consolas", 12), width=16,
                     bg="#1a1a1a", fg="#00ff88", insertbackground="#00ff88",
                     relief=tk.SUNKEN, bd=2)
        e.pack()
        e.focus_set()
        e.selection_range(0, tk.END)

        def submit():
            try:
                new_val = float(v.get().strip())
                if new_val < 0:
                    raise ValueError
            except ValueError:
                self._show_toast("请输入有效数字（>=0）")
                return

            if field == x_key:
                for i, dp in enumerate(self.data_points):
                    if i != row_idx and abs(dp[x_key] - new_val) < 1e-9:
                        self._show_toast("已存在相同 {}，不能重复".format(x_name))
                        return

            self.data_points[row_idx][field] = new_val
            # 编辑后强制按报告结论回归标准结果，保证可正确计算标准数据
            if self.is_distance_experiment:
                d_now = float(self.data_points[row_idx]["d"])
                u_std, i_std = self._measure_ui_from_distance_data(d_now)
                self.data_points[row_idx]["U"] = float(u_std)
                self.data_points[row_idx]["I"] = float(i_std)
                self.data_points[row_idx]["P"] = float(self._lookup_distance_light_intensity(d_now))
            else:
                r_now = float(self.data_points[row_idx]["R"])
                u_std, i_std = self._infer_standard_iv_from_report(r_now)
                self.data_points[row_idx]["U"] = float(u_std)
                self.data_points[row_idx]["I"] = float(i_std)
                self.data_points[row_idx]["P"] = float(u_std) * float(i_std)

            dp = self.data_points[row_idx]
            if self.is_distance_experiment:
                self.tree.item(item_id, values=(
                    row_idx + 1,
                    "{:g}".format(dp[x_key]),
                    "{:.3f}".format(dp["P"]),
                    "{:.3f}".format(dp["U"]),
                    "{:.3f}".format(dp["I"]),
                ))
            else:
                self.tree.item(item_id, values=(
                    row_idx + 1,
                    "{:g}".format(dp[x_key]),
                    "{:.3f}".format(dp["U"]),
                    "{:.3f}".format(dp["I"]),
                    "{:.3f}".format(dp["P"]),
                ))

            if len(self.data_points) >= 10:
                self._draw_plot()
            dlg.destroy()

        btns = tk.Frame(dlg, bg=BG)
        btns.pack(fill=tk.X, padx=24, pady=16)
        ttk.Button(btns, text="确认", style="Primary.TButton",
                   command=submit).pack(side=tk.LEFT, expand=True, padx=4)
        ttk.Button(btns, text="取消", style="Assist.TButton",
                   command=dlg.destroy).pack(side=tk.LEFT, expand=True, padx=4)
        dlg.bind("<Return>", lambda _e: submit())
        dlg.bind("<Escape>", lambda _e: dlg.destroy())

    def _show_toast(self, msg):
        toast = tk.Toplevel(self.frame.winfo_toplevel())
        toast.overrideredirect(True)
        toast.configure(bg="#ff4444")
        x = self.frame.winfo_rootx() + self.frame.winfo_width() // 2 - 150
        y = self.frame.winfo_rooty() + 80
        toast.geometry("300x36+{}+{}".format(x, y))
        tk.Label(toast, text=msg, bg="#ff4444", fg="#fff",
                 font=("Microsoft YaHei", 10, "bold")).pack(expand=True, fill=tk.BOTH)
        toast.after(2500, toast.destroy)

    def _clear_data(self):
        self.data_points.clear()
        self.exp2_partial_points = {}
        self.last_auto_record_r = None
        self._refresh_table_view()
        if self.ax1 is not None and self.canvas1 is not None:
            self.ax1.clear()
            self._style_ax(self.ax1, title="伏安特性曲线", xlabel="U (V)", ylabel="I (mA)")
            self.canvas1.draw_idle()
        if self.ax2 is not None and self.canvas2 is not None:
            self.ax2.clear()
            self._style_ax(self.ax2, title="功率输出曲线（P-V）", xlabel="U (V)", ylabel="P (mW)")
            self.canvas2.draw_idle()

    def _load_standard_data(self):
        self._clear_data()
        if self.is_distance_experiment:
            data = get_experimental_data()["distance_data"]
            for d, e, voc, isc in zip(data["d_cm"], data["E_wm2"], data["Voc_V"], data["Isc_mA"]):
                point = {"d": float(d), "P": float(e), "U": float(voc), "I": float(isc)}
                self.data_points.append(point)
            self._refresh_table_view()
            self._show_toast("实验二已填充 {} 组标准数据（请手动点击查看曲线/分析）".format(len(self.data_points)))
            return

        data = get_experimental_data()["iv_data"]
        for r, u, i in zip(data["R"], data["U"], data["I"]):
            # 标准数据中含短路和近开路点，保留用于分析。
            p = u * i
            point = {"R": float(r), "U": float(u), "I": float(i), "P": float(p)}
            self.data_points.append(point)
        self._refresh_table_view()
        ok, msg = self._validate_exp1_data()
        if not ok:
            self._show_toast("标准数据校验失败: " + msg)
            return
        self._show_toast("实验一已填充 {} 组标准数据并通过校验（请手动点击查看曲线/分析）".format(len(self.data_points)))

    def _validate_exp1_data(self):
        if not self.data_points:
            return False, "无数据"
        for dp in self.data_points:
            if dp["R"] < 0 or dp["U"] < 0 or dp["I"] < 0:
                return False, "存在负值"
            if dp["U"] > 10 or dp["I"] > 200:
                return False, "超出量程"
            if abs(dp["P"] - dp["U"] * dp["I"]) > 1e-9:
                return False, "功率计算错误"
        return True, "ok"

    # ── 绘图（≥10 组才绘图）──

    def _draw_plot(self, auto_open_window=True):
        min_points = 2 if self.is_distance_experiment else 10
        if len(self.data_points) < min_points:
            self._show_toast("需要至少 {} 组数据才能绘图！当前 {} 组".format(min_points, len(self.data_points)))
            return
        if (self.curve_win is None or not self.curve_win.winfo_exists()) and auto_open_window:
            self._open_curve_window()
        if self.ax1 is None or self.ax2 is None or self.canvas1 is None or self.canvas2 is None:
            return

        Us = [d["U"] for d in self.data_points]
        Is = [d["I"] for d in self.data_points]
        Ps = [d["P"] for d in self.data_points]

        def _dense_xy(xs, ys, n=200):
            if len(xs) < 2:
                return list(xs), list(ys)
            x_arr = np.array(xs, dtype=float)
            y_arr = np.array(ys, dtype=float)
            xmin, xmax = float(x_arr.min()), float(x_arr.max())
            xmid = xmin + 0.5 * (xmax - xmin)
            n1 = max(24, int(n * 0.35))
            n2 = max(36, n - n1)
            x_dense_1 = np.linspace(xmin, xmid, n1, endpoint=False)
            x_dense_2 = np.linspace(xmid, xmax, n2)
            x_dense = np.concatenate((x_dense_1, x_dense_2))
            y_dense = np.interp(x_dense, x_arr, y_arr)
            return x_dense, y_dense

        def _enforce_nonincreasing(vals):
            if not vals:
                return vals
            out = [float(vals[0])]
            for v in vals[1:]:
                out.append(min(out[-1], float(v)))
            return out

        def _extra_points_right(xs, ys, extra_n=12):
            if len(xs) < 2 or extra_n <= 0:
                return np.array([]), np.array([])
            x_arr = np.array(xs, dtype=float)
            y_arr = np.array(ys, dtype=float)
            xmin, xmax = float(x_arr.min()), float(x_arr.max())
            # 两段补点：中段少量，末端密集（解决最右下角稀疏）
            x1_start = xmin + 0.88 * (xmax - xmin)
            x1_end = xmin + 0.945 * (xmax - xmin)
            x2_start = x1_end
            x2_end = xmax
            n_mid = max(3, int(extra_n * 0.35))
            n_tail = max(6, extra_n - n_mid)
            t1 = np.linspace(0.0, 1.0, n_mid) ** 1.3
            t2 = np.linspace(0.0, 1.0, n_tail) ** 3.2
            x_mid = x1_start + (x1_end - x1_start) * t1
            x_tail = x2_start + (x2_end - x2_start) * t2
            x_extra = np.concatenate((x_mid, x_tail))
            y_extra = np.interp(x_extra, x_arr, y_arr)
            return x_extra, y_extra

        def _merge_points(xs, ys, x_extra, y_extra):
            if len(x_extra) == 0:
                return list(xs), list(ys)
            x_all = np.concatenate((np.array(xs, dtype=float), np.array(x_extra, dtype=float)))
            y_all = np.concatenate((np.array(ys, dtype=float), np.array(y_extra, dtype=float)))
            order = np.argsort(x_all)
            return [float(x_all[i]) for i in order], [float(y_all[i]) for i in order]

        def _fit_smooth(xs, ys, n=320):
            if len(xs) < 3:
                return np.array(xs, dtype=float), np.array(ys, dtype=float)
            x_arr = np.array(xs, dtype=float)
            y_arr = np.array(ys, dtype=float)
            uniq_x, uniq_idx = np.unique(x_arr, return_index=True)
            uniq_y = y_arr[uniq_idx]
            if len(uniq_x) < 3:
                return uniq_x, uniq_y
            x_dense = np.linspace(float(uniq_x.min()), float(uniq_x.max()), n)
            try:
                f = PchipInterpolator(uniq_x, uniq_y)
                y_dense = f(x_dense)
            except Exception:
                y_dense = np.interp(x_dense, uniq_x, uniq_y)
            return x_dense, y_dense

        def _smooth_display(xs, ys, n=280):
            if len(xs) < 3:
                return np.array(xs, dtype=float), np.array(ys, dtype=float)
            x_arr = np.array(xs, dtype=float)
            y_arr = np.array(ys, dtype=float)
            uniq_x, uniq_idx = np.unique(x_arr, return_index=True)
            uniq_y = y_arr[uniq_idx]
            if len(uniq_x) < 3:
                return uniq_x, uniq_y
            x_dense = np.linspace(float(uniq_x.min()), float(uniq_x.max()), n)
            try:
                f = PchipInterpolator(uniq_x, uniq_y)
                y_dense = f(x_dense)
            except Exception:
                y_dense = np.interp(x_dense, uniq_x, uniq_y)
            return x_dense, y_dense

        # ── 图1：伏安特性曲线 (I vs U) ──
        self.ax1.clear()
        self._style_ax(self.ax1, title="伏安特性曲线", xlabel="U (V)", ylabel="I (mA)")
        order_u = np.argsort(Us)
        Us_sorted = [Us[i] for i in order_u]
        Is_sorted = [Is[i] for i in order_u]
        # 实验报告中的 I-V 曲线应随电压单调下降，这里做单调约束，避免下沉后回升。
        Is_curve = _enforce_nonincreasing(Is_sorted) if not self.is_distance_experiment else Is_sorted
        u_dense, i_dense = _dense_xy(Us_sorted, Is_curve)
        if self.curve_fit_mode_1:
            x_fit, y_fit = _fit_smooth(Us_sorted, Is_curve)
            self.ax1.plot(x_fit, y_fit, "-", color=ACCENT, linewidth=2.4, label="I-U 拟合")
        else:
            x_show, y_show = _smooth_display(Us_sorted, Is_curve)
            self.ax1.plot(x_show, y_show, "-", color=ACCENT, linewidth=2.2, label="I-U")
        u_mark = Us_sorted
        i_mark = Is_curve if not self.is_distance_experiment else Is_sorted
        if not self.is_distance_experiment:
            ux, ix = _extra_points_right(Us_sorted, Is_curve, extra_n=12)
            u_mark, i_mark = _merge_points(Us_sorted, Is_curve, ux, ix)
        if not self.curve_fit_mode_1:
            self.ax1.plot(u_mark, i_mark, "o", color=ACCENT, markersize=4)
        self.ax1.legend(loc="upper right", fontsize=8, framealpha=0.4,
                        labelcolor=FG, prop=plt_font)
        self.fig1.tight_layout()
        self.canvas1.draw_idle()

        # ── 图2：功率输出曲线 (P vs U) ──
        self.ax2.clear()
        self._style_ax(self.ax2, title="功率输出曲线（P-V）", xlabel="U (V)", ylabel="P (mW)")
        order_u2 = np.argsort(Us)
        Us_sorted2 = [Us[i] for i in order_u2]
        Ps_sorted = [Ps[i] for i in order_u2]
        u2_dense, p_dense = _dense_xy(Us_sorted2, Ps_sorted)
        if self.curve_fit_mode_2:
            x_fit2, y_fit2 = _fit_smooth(Us_sorted2, Ps_sorted)
            self.ax2.plot(x_fit2, y_fit2, "-", color=ACCENT2, linewidth=2.4, label="P-V 拟合")
        else:
            x_show2, y_show2 = _smooth_display(Us_sorted2, Ps_sorted)
            self.ax2.plot(x_show2, y_show2, "-", color=ACCENT2, linewidth=2.2, label="P-V")
        u2_mark = Us_sorted2
        p_mark = Ps_sorted
        if not self.is_distance_experiment:
            ux2, px2 = _extra_points_right(Us_sorted2, Ps_sorted, extra_n=12)
            u2_mark, p_mark = _merge_points(Us_sorted2, Ps_sorted, ux2, px2)
        if not self.curve_fit_mode_2:
            self.ax2.plot(u2_mark, p_mark, "s", color=ACCENT2, markersize=4)
        self.ax2.legend(loc="upper right", fontsize=8, framealpha=0.4,
                        labelcolor=FG, prop=plt_font)
        self.fig2.tight_layout()
        self.canvas2.draw_idle()

        if self.is_distance_experiment:
            intensities = [d["P"] for d in self.data_points]
            self.ax1.clear()
            self._style_ax(self.ax1, title="开路电压-光强关系曲线", xlabel="I (W/m²)", ylabel="VOC (V)")
            order_i = np.argsort(intensities)
            i_sorted = [intensities[i] for i in order_i]
            voc_sorted = [Us[i] for i in order_i]
            i_dense, voc_dense = _dense_xy(i_sorted, voc_sorted)
            if self.curve_fit_mode_1:
                x_fit, y_fit = _fit_smooth(i_sorted, voc_sorted)
                self.ax1.plot(x_fit, y_fit, "-", color=ACCENT, linewidth=2.4, label="VOC-I 拟合")
            else:
                x_show, y_show = _smooth_display(i_sorted, voc_sorted)
                self.ax1.plot(x_show, y_show, "-", color=ACCENT, linewidth=2.2, label="VOC-I")
                self.ax1.plot(i_sorted, voc_sorted, "o", color=ACCENT, markersize=4)
                if len(i_sorted) >= 2:
                    n_extra = len(i_sorted) + 6
                    i_extra = np.linspace(float(i_sorted[0]), float(i_sorted[-1]), n_extra)
                    voc_extra = np.interp(i_extra, x_show, y_show)
                    self.ax1.plot(i_extra, voc_extra, "o", color=ACCENT, markersize=2.4, alpha=0.6)
            self.ax1.legend(loc="upper right", fontsize=8, framealpha=0.4,
                            labelcolor=FG, prop=plt_font)
            self.fig1.tight_layout()
            self.canvas1.draw_idle()

            self.ax2.clear()
            self._style_ax(self.ax2, title="短路电流-光强关系曲线", xlabel="I (W/m²)", ylabel="ISC (mA)")
            isc_sorted = [Is[i] for i in order_i]
            i2_dense, isc_dense = _dense_xy(i_sorted, isc_sorted)
            if self.curve_fit_mode_2:
                x_fit2, y_fit2 = _fit_smooth(i_sorted, isc_sorted)
                self.ax2.plot(x_fit2, y_fit2, "-", color=ACCENT2, linewidth=2.4, label="ISC-I 拟合")
            else:
                x_show2, y_show2 = _smooth_display(i_sorted, isc_sorted)
                self.ax2.plot(x_show2, y_show2, "-", color=ACCENT2, linewidth=2.2, label="ISC-I")
                self.ax2.plot(i_sorted, isc_sorted, "s", color=ACCENT2, markersize=4)
                if len(i_sorted) >= 2:
                    n_extra = len(i_sorted) + 6
                    i_extra = np.linspace(float(i_sorted[0]), float(i_sorted[-1]), n_extra)
                    isc_extra = np.interp(i_extra, i_sorted, isc_sorted)
                    self.ax2.plot(i_extra, isc_extra, "s", color=ACCENT2, markersize=2.2, alpha=0.6)
            self.ax2.legend(loc="upper right", fontsize=8, framealpha=0.4,
                            labelcolor=FG, prop=plt_font)
            self.fig2.tight_layout()
            self.canvas2.draw_idle()

        if self.curve_btn1 is not None and self.curve_btn1.winfo_exists():
            self.curve_btn1.config(text="恢复原始图像" if self.curve_fit_mode_1 else "展示拟合曲线")
        if self.curve_btn2 is not None and self.curve_btn2.winfo_exists():
            self.curve_btn2.config(text="恢复原始图像" if self.curve_fit_mode_2 else "展示拟合曲线")

    # ── 数据分析弹窗 ──

    def _show_analysis(self):
        if len(self.data_points) < 10:
            self._show_toast("需要至少 10 组数据才能分析！当前 {} 组".format(len(self.data_points)))
            return

        best_idx = self._best_point_index()
        best = self.data_points[best_idx]
        max_p = best["P"]

        R0 = best["d"] if self.is_distance_experiment else best["R"]
        Um = best["U"]
        Im = best["I"]
        if self.is_distance_experiment:
            win = tk.Toplevel(self.frame.winfo_toplevel())
            win.title("数据分析结果")
            win.configure(bg=BG)
            win.geometry("520x360")
            win.resizable(False, False)

            tk.Label(win, text="数据分析结果", bg=ACCENT, fg="#fff",
                     font=("Microsoft YaHei", 14, "bold")).pack(fill=tk.X, ipady=8)
            tk.Label(win, text="", bg=BG).pack(pady=(8, 6))
            table = tk.Frame(win, bg=PANEL_BG)
            table.pack(fill=tk.X, padx=30)
            results = [
                ("最佳距离 d*", "{:g} cm".format(R0)),
                ("对应电压 U", "{:.3f} V".format(Um)),
                ("对应电流 I", "{:.3f} mA".format(Im)),
                ("最大输出功率 Pmax", "{:.3f} mW".format(max_p)),
                ("当前负载电阻", "{:g} Ω".format(self.r_value)),
            ]
            for i, (label, value) in enumerate(results):
                bg_c = "#1a2a3e" if i % 2 == 0 else "#16213e"
                row = tk.Frame(table, bg=bg_c)
                row.pack(fill=tk.X, ipady=6)
                tk.Label(row, text="  " + label, bg=bg_c, fg="#ffffff",
                         font=("Microsoft YaHei", 11), anchor="w").pack(side=tk.LEFT)
                tk.Label(row, text=value + "  ", bg=bg_c, fg="#ffffff",
                         font=("Consolas", 13, "bold"), anchor="e").pack(side=tk.RIGHT)
            tk.Label(win, text="", bg=BG).pack(pady=8)
            ttk.Button(win, text="关闭", style="Assist.TButton",
                       command=win.destroy).pack(ipady=2, padx=20)
            return

        # 仅使用已记录实测点做分析，不做任何拟合/外推：
        # Uoc_exp: 取最大电阻实测点电压（最接近开路）
        # Isc_exp: 取最小电阻实测点电流（最接近短路）
        nonzero_r = [d for d in self.data_points if d["R"] > 0]
        if nonzero_r:
            max_r_point = max(nonzero_r, key=lambda d: d["R"])
        else:
            max_r_point = max(self.data_points, key=lambda d: d["U"])
        min_r_point = min(self.data_points, key=lambda d: d["R"])
        Uoc_exp = max_r_point["U"]
        Isc_exp = min_r_point["I"]
        # FF = Im*Um / (Isc*Uoc)，按百分数显示
        FF = (Im * Um) / (Isc_exp * Uoc_exp) if (Uoc_exp * Isc_exp) > 0 else 0
        FF_percent = FF * 100.0

        win = tk.Toplevel(self.frame.winfo_toplevel())
        win.title("数据分析结果")
        win.configure(bg=BG)
        win.geometry("520x440")
        win.resizable(False, False)

        tk.Label(win, text="数据分析结果", bg=ACCENT, fg="#fff",
                 font=("Microsoft YaHei", 14, "bold")).pack(fill=tk.X, ipady=8)
        tk.Label(win, text="", bg=BG).pack(pady=(8, 6))

        table = tk.Frame(win, bg=PANEL_BG)
        table.pack(fill=tk.X, padx=30)

        results = [
            ("最佳负载电阻 R₀", "{:g} Ω".format(R0)),
            ("最佳工作电压 Uₘ", "{:.3f} V".format(Um)),
            ("最佳工作电流 Iₘ", "{:.3f} mA".format(Im)),
            ("开路电压 Uoc", "{:.3f} V".format(Uoc_exp)),
            ("短路电流 Isc", "{:.3f} mA".format(Isc_exp)),
            ("填充因子 F·F", "{:.2f} %".format(FF_percent)),
        ]

        for i, (label, value) in enumerate(results):
            bg_c = "#1a2a3e" if i % 2 == 0 else "#16213e"
            row = tk.Frame(table, bg=bg_c)
            row.pack(fill=tk.X, ipady=6)
            tk.Label(row, text="  " + label, bg=bg_c, fg="#ffffff",
                     font=("Microsoft YaHei", 11), anchor="w").pack(side=tk.LEFT)
            tk.Label(row, text=value + "  ", bg=bg_c, fg="#ffffff",
                     font=("Consolas", 13, "bold"), anchor="e").pack(side=tk.RIGHT)

        tk.Label(win, text="", bg=BG).pack(pady=8)
        ttk.Button(win, text="关闭", style="Assist.TButton",
                   command=win.destroy).pack(ipady=2, padx=20)

    def _get_current_params(self):
        p = dict(self.params)
        base_Iph = DEFAULT_PARAMS["single"]["Iph"]
        p["Iph"] = base_Iph * (self.light_intensity / 1000.0)
        return p

    def export_data_excel(self):
        if not self.data_points:
            messagebox.showinfo("提示", f"{self.experiment_name}暂无可导出数据")
            return
        path = filedialog.asksaveasfilename(
            title=f"导出{self.experiment_name}数据",
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx")],
            initialfile=f"{self.experiment_name}_数据.xlsx",
        )
        if not path:
            return
        try:
            from openpyxl import Workbook
        except Exception:
            messagebox.showerror("错误", "缺少 openpyxl，请先安装：pip install openpyxl")
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "数据"
        if self.is_distance_experiment:
            headers = ["序号", "d(cm)", "I(W/m²)", "Voc(V)", "Isc(mA)"]
            ws.append(headers)
            rows = sorted(self.data_points, key=lambda d: d["d"])
            for i, dp in enumerate(rows, start=1):
                ws.append([i, dp["d"], dp["P"], dp["U"], dp["I"]])
        else:
            headers = ["序号", "R(Ω)", "U(V)", "I(mA)", "P(mW)"]
            ws.append(headers)
            rows = sorted(self.data_points, key=lambda d: d["R"])
            for i, dp in enumerate(rows, start=1):
                ws.append([i, dp["R"], dp["U"], dp["I"], dp["P"]])
            # 实验一额外附加数据分析结果
            if len(self.data_points) >= 10:
                ana = wb.create_sheet("数据分析")
                best_idx = self._best_point_index()
                best = self.data_points[best_idx]
                nonzero_r = [d for d in self.data_points if d["R"] > 0]
                max_r_point = max(nonzero_r, key=lambda d: d["R"]) if nonzero_r else max(self.data_points, key=lambda d: d["U"])
                min_r_point = min(self.data_points, key=lambda d: d["R"])
                Uoc_exp = max_r_point["U"]
                Isc_exp = min_r_point["I"]
                FF = (best["I"] * best["U"]) / (Isc_exp * Uoc_exp) if (Uoc_exp * Isc_exp) > 0 else 0.0
                items = [
                    ("最佳负载电阻 R0 (Ω)", best["R"]),
                    ("最佳工作电压 Um (V)", best["U"]),
                    ("最佳工作电流 Im (mA)", best["I"]),
                    ("最大功率 Pmax (mW)", best["P"]),
                    ("开路电压 Uoc (V)", Uoc_exp),
                    ("短路电流 Isc (mA)", Isc_exp),
                    ("填充因子 FF (%)", FF * 100.0),
                ]
                ana.append(["指标", "数值"])
                for k, v in items:
                    ana.append([k, float(v)])
                ana.column_dimensions["A"].width = 28
                ana.column_dimensions["B"].width = 16
        for col in ("A", "B", "C", "D", "E"):
            ws.column_dimensions[col].width = 16
        wb.save(path)
        messagebox.showinfo("完成", f"已导出：\n{path}")

    def export_chart(self):
        min_points = 2 if self.is_distance_experiment else 10
        if len(self.data_points) < min_points:
            messagebox.showinfo("提示", f"{self.experiment_name}暂无可导出图表（至少需要 {min_points} 组数据）")
            return
        base = filedialog.asksaveasfilename(
            title=f"导出{self.experiment_name}图表",
            defaultextension=".png",
            filetypes=[("PNG 图片", "*.png")],
            initialfile=f"{self.experiment_name}_图表.png",
        )
        if not base:
            return
        root, ext = os.path.splitext(base)
        ext = ext or ".png"
        p1 = f"{root}_图1{ext}"
        p2 = f"{root}_图2{ext}"
        p3 = f"{root}_合并{ext}"

        def _smooth_xy(xs, ys, n=260):
            if len(xs) < 3:
                return np.array(xs), np.array(ys)
            x = np.array(xs, dtype=float)
            y = np.array(ys, dtype=float)
            ux, ui = np.unique(x, return_index=True)
            uy = y[ui]
            xd = np.linspace(float(ux.min()), float(ux.max()), n)
            try:
                f = PchipInterpolator(ux, uy)
                yd = f(xd)
            except Exception:
                yd = np.interp(xd, ux, uy)
            return xd, yd

        def _plot_one(ax, idx):
            if self.is_distance_experiment:
                pts = sorted(self.data_points, key=lambda d: d["P"])
                xs = [d["P"] for d in pts]
                y1 = [d["U"] for d in pts]
                y2 = [d["I"] for d in pts]
                if idx == 1:
                    xd, yd = _smooth_xy(xs, y1)
                    ax.plot(xd, yd, "-", color=ACCENT, linewidth=2.2)
                    ax.plot(xs, y1, "o", color=ACCENT, markersize=4)
                    self._style_ax(ax, title="开路电压-光强关系曲线", xlabel="I (W/m²)", ylabel="VOC (V)")
                else:
                    xd, yd = _smooth_xy(xs, y2)
                    ax.plot(xd, yd, "-", color=ACCENT2, linewidth=2.2)
                    ax.plot(xs, y2, "s", color=ACCENT2, markersize=4)
                    self._style_ax(ax, title="短路电流-光强关系曲线", xlabel="I (W/m²)", ylabel="ISC (mA)")
            else:
                Us = [d["U"] for d in self.data_points]
                Is = [d["I"] for d in self.data_points]
                Ps = [d["P"] for d in self.data_points]
                order = np.argsort(Us)
                x = [Us[i] for i in order]
                if idx == 1:
                    y = [Is[i] for i in order]
                    xd, yd = _smooth_xy(x, y)
                    ax.plot(xd, yd, "-", color=ACCENT, linewidth=2.2)
                    ax.plot(x, y, "o", color=ACCENT, markersize=4)
                    self._style_ax(ax, title="伏安特性曲线", xlabel="U (V)", ylabel="I (mA)")
                else:
                    y = [Ps[i] for i in order]
                    xd, yd = _smooth_xy(x, y)
                    ax.plot(xd, yd, "-", color=ACCENT2, linewidth=2.2)
                    ax.plot(x, y, "s", color=ACCENT2, markersize=4)
                    self._style_ax(ax, title="功率输出曲线（P-V）", xlabel="U (V)", ylabel="P (mW)")

        fig_a = Figure(figsize=(6, 4), dpi=150, facecolor=BG)
        ax_a = fig_a.add_subplot(111)
        _plot_one(ax_a, 1)
        fig_a.tight_layout()
        fig_a.savefig(p1, dpi=150, bbox_inches="tight")

        fig_b = Figure(figsize=(6, 4), dpi=150, facecolor=BG)
        ax_b = fig_b.add_subplot(111)
        _plot_one(ax_b, 2)
        fig_b.tight_layout()
        fig_b.savefig(p2, dpi=150, bbox_inches="tight")

        fig_c = Figure(figsize=(12, 4), dpi=150, facecolor=BG)
        ax_c1 = fig_c.add_subplot(121)
        ax_c2 = fig_c.add_subplot(122)
        _plot_one(ax_c1, 1)
        _plot_one(ax_c2, 2)
        fig_c.tight_layout()
        fig_c.savefig(p3, dpi=150, bbox_inches="tight")
        messagebox.showinfo("完成", f"已导出：\n{p1}\n{p2}\n{p3}")


# ═══════════════════════════════════════════════
#  实验二、三：自动计算版本
# ═══════════════════════════════════════════════

class ExperimentTab:
    """实验二/三：按指导书流程自动计算数据。"""

    def __init__(self, parent_frame, title, experiment_type="distance"):
        self.frame = parent_frame
        self.title = title
        self.experiment_type = experiment_type
        self.params = {k: v for k, v in DEFAULT_PARAMS["single"].items()}
        self.data_points = []
        self.distance_cm = 30.0
        self.lamp_power_w = 100.0
        self.light_intensity = calc_light_intensity(self.distance_cm)
        self.fixed_R = 100.0
        self.ax2 = None
        self._build()

    def _build(self):
        body = tk.Frame(self.frame, bg=BG)
        body.pack(fill=tk.BOTH, expand=True)
        mid = tk.Frame(body, bg=PANEL_BG, width=420)
        mid.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 6))
        mid.pack_propagate(False)
        right = tk.Frame(body, bg=BG)
        right.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self._build_control(mid)
        self._build_right(right)

        bottom = tk.Frame(self.frame, bg=BG)
        bottom.pack(fill=tk.X, pady=(6, 0))
        ttk.Button(bottom, text="数据分析", style="Info.TButton",
                   command=self._draw_plot).pack(fill=tk.X, padx=10, ipady=2)

    def _build_control(self, parent):
        tk.Label(parent, text="电池类型", bg=PANEL_BG, fg=FG,
                 font=("Microsoft YaHei", 10, "bold")).pack(anchor="w", padx=10, pady=(10, 2))
        tf = tk.Frame(parent, bg=PANEL_BG)
        tf.pack(fill=tk.X, padx=10)
        self.cell_type_var = tk.StringVar(value="single")
        tk.Label(tf, text="单晶硅", bg=PANEL_BG, fg=FG,
                 font=("Microsoft YaHei", 9)).pack(side=tk.LEFT, padx=5)
        ttk.Separator(parent, orient="horizontal").pack(fill=tk.X, padx=10, pady=6)

        if self.experiment_type == "distance":
            self._build_distance_control(parent)
        elif self.experiment_type == "light":
            self._build_light_control(parent)

        ttk.Separator(parent, orient="horizontal").pack(fill=tk.X, padx=10, pady=6)

        tk.Label(parent, text="━━━ 实时读数 ━━━", bg=PANEL_BG, fg="#888",
                 font=("Consolas", 9)).pack(pady=(4, 6))
        v_text = "Voc = 0.000 V" if self.experiment_type == "distance" else "U = 0.000 V"
        i_text = "Isc = 0.000 mA" if self.experiment_type == "distance" else "I = 0.000 mA"
        p_text = "E = 0 W/m²" if self.experiment_type == "distance" else "P = 0.000 mW"
        self.reading_v = tk.Label(parent, text=v_text, bg="#0a0a0a",
                                  fg="#00ff88", font=("Consolas", 16, "bold"),
                                  relief=tk.SUNKEN, bd=2, padx=8, pady=4)
        self.reading_v.pack(fill=tk.X, padx=10, pady=2)
        self.reading_i = tk.Label(parent, text=i_text, bg="#0a0a0a",
                                  fg="#00ff88", font=("Consolas", 16, "bold"),
                                  relief=tk.SUNKEN, bd=2, padx=8, pady=4)
        self.reading_i.pack(fill=tk.X, padx=10, pady=2)
        self.reading_p = tk.Label(parent, text=p_text, bg="#0a0a0a",
                                  fg="#ffcc00", font=("Consolas", 14, "bold"),
                                  relief=tk.SUNKEN, bd=2, padx=8, pady=4)
        self.reading_p.pack(fill=tk.X, padx=10, pady=2)
        self._update_reading()

        ttk.Separator(parent, orient="horizontal").pack(fill=tk.X, padx=10, pady=6)

        ttk.Button(parent, text="● 记录当前数据点", style="Primary.TButton",
                   command=self._record_point).pack(fill=tk.X, padx=10, pady=4, ipady=2)

        btn_frame = tk.Frame(parent, bg=PANEL_BG)
        btn_frame.pack(fill=tk.X, padx=10, pady=4)
        ttk.Button(btn_frame, text="清除数据", style="Assist.TButton",
                   command=self._clear_data).pack(fill=tk.X, padx=2)

    def _build_distance_control(self, parent):
        tk.Label(parent, text="━━━ 光源距离 ━━━", bg=PANEL_BG, fg=ACCENT,
                 font=("Microsoft YaHei", 10, "bold")).pack(anchor="w", padx=10, pady=(0, 2))
        self.dist_var = tk.DoubleVar(value=30.0)
        tk.Scale(parent, from_=5, to=100, orient=tk.HORIZONTAL,
                 variable=self.dist_var, bg=PANEL_BG, fg=FG,
                 troughcolor="#333", highlightthickness=0,
                 resolution=1, showvalue=True, length=200,
                 font=("Consolas", 9),
                 command=self._on_distance_change).pack(fill=tk.X, padx=10)
        self.dist_label = tk.Label(parent, text="d = 30 cm", bg=PANEL_BG, fg=FG,
                                   font=("Consolas", 12, "bold"))
        self.dist_label.pack(pady=4)
        info = tk.Frame(parent, bg=PANEL_BG)
        info.pack(fill=tk.X, padx=10, pady=4)
        tk.Label(info, text="固定条件:", bg=PANEL_BG, fg="#888",
                 font=("Microsoft YaHei", 9)).pack(anchor="w")
        tk.Label(info, text="  光源功率: 100 W", bg=PANEL_BG, fg=FG,
                 font=("Microsoft YaHei", 9)).pack(anchor="w")
        tk.Label(info, text="  测量: 开路电压 Voc / 短路电流 Isc", bg=PANEL_BG, fg=FG,
                 font=("Microsoft YaHei", 9)).pack(anchor="w")

    def _build_light_control(self, parent):
        tk.Label(parent, text="━━━ 光源功率 ━━━", bg=PANEL_BG, fg=ACCENT,
                 font=("Microsoft YaHei", 10, "bold")).pack(anchor="w", padx=10, pady=(0, 2))
        self.power_var = tk.DoubleVar(value=100.0)
        tk.Scale(parent, from_=10, to=300, orient=tk.HORIZONTAL,
                 variable=self.power_var, bg=PANEL_BG, fg=FG,
                 troughcolor="#333", highlightthickness=0,
                 resolution=5, showvalue=True, length=200,
                 font=("Consolas", 9),
                 command=self._on_power_change).pack(fill=tk.X, padx=10)
        self.power_label = tk.Label(parent, text="P_lamp = 100 W", bg=PANEL_BG, fg=FG,
                                    font=("Consolas", 12, "bold"))
        self.power_label.pack(pady=4)
        self.intensity_label = tk.Label(parent, text="E = {:.0f} W/m²".format(self.light_intensity), bg=PANEL_BG,
                                        fg="#aaa", font=("Consolas", 10))
        self.intensity_label.pack()
        info = tk.Frame(parent, bg=PANEL_BG)
        info.pack(fill=tk.X, padx=10, pady=4)
        tk.Label(info, text="固定条件:", bg=PANEL_BG, fg="#888",
                 font=("Microsoft YaHei", 9)).pack(anchor="w")
        tk.Label(info, text="  光源-板距离: 30 cm", bg=PANEL_BG, fg=FG,
                 font=("Microsoft YaHei", 9)).pack(anchor="w")
        tk.Label(info, text="  负载电阻: 100 Ω", bg=PANEL_BG, fg=FG,
                 font=("Microsoft YaHei", 9)).pack(anchor="w")

    def _build_right(self, parent):
        chart_frame = tk.Frame(parent, bg=BG)
        chart_frame.pack(fill=tk.BOTH, expand=True)
        self.fig = Figure(figsize=(6, 4), dpi=100, facecolor=BG)
        self.ax = self.fig.add_subplot(111)
        self._style_ax(self.ax)
        self.canvas = FigureCanvasTkAgg(self.fig, master=chart_frame)
        self.canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

        table_frame = tk.Frame(parent, bg=PANEL_BG, height=180)
        table_frame.pack(fill=tk.X, padx=0, pady=(4, 0))
        table_frame.pack_propagate(False)

        hdr = tk.Frame(table_frame, bg=PANEL_BG)
        hdr.pack(fill=tk.X, padx=8, pady=(4, 2))
        tk.Label(hdr, text="实验数据记录", bg=PANEL_BG, fg=ACCENT,
                 font=("Microsoft YaHei", 10, "bold")).pack(side=tk.LEFT)
        ttk.Button(hdr, text="删除选中行", style="Assist.TButton",
                   command=self._delete_selected).pack(side=tk.RIGHT, padx=4)

        if self.experiment_type == "distance":
            cols = ("序号", "d (cm)", "E (W/m²)", "Voc (V)", "Isc (mA)")
        else:
            cols = ("序号", "E (W/m²)", "U (V)", "I (mA)", "P (mW)")
        self.tree = ttk.Treeview(table_frame, columns=cols, show="headings", height=6)
        for c in cols:
            self.tree.heading(c, text=c)
            w = 50 if c == "序号" else 100
            self.tree.column(c, width=w, anchor="center")
        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(8, 0))
        vsb.pack(side=tk.RIGHT, fill=tk.Y, padx=(0, 8))

    def _style_ax(self, ax, title="", xlabel="", ylabel=""):
        ax.set_facecolor("#0f0f23")
        ax.set_title(title, fontproperties=plt_font, color=FG, fontsize=12, pad=8)
        ax.set_xlabel(xlabel, fontproperties=plt_font, color=FG, fontsize=10)
        ax.set_ylabel(ylabel, fontproperties=plt_font, color=FG, fontsize=10)
        ax.tick_params(colors="#888", labelsize=8)
        for spine in ax.spines.values():
            spine.set_color(GRID_COLOR)
        ax.grid(True, alpha=0.2, color=GRID_COLOR)

    def _on_type_change(self):
        self.params = {k: v for k, v in DEFAULT_PARAMS["single"].items()}
        self._update_reading()

    def _on_distance_change(self, val):
        d = float(val)
        self.distance_cm = d
        self.light_intensity = calc_light_intensity(d) * (self.lamp_power_w / 100.0)
        self.dist_label.config(text="d = {:.0f} cm".format(d))
        self._update_reading()

    def _on_power_change(self, val):
        p = float(val)
        self.lamp_power_w = p
        self.light_intensity = calc_light_intensity(self.distance_cm) * (p / 100.0)
        self.power_label.config(text="P_lamp = {:.0f} W".format(p))
        self.intensity_label.config(text="E = {:.0f} W/m²".format(self.light_intensity))
        self._update_reading()

    def _get_current_params(self):
        p = dict(self.params)
        base_Iph = DEFAULT_PARAMS["single"]["Iph"]
        p["Iph"] = base_Iph * (self.light_intensity / 1000.0)
        return p

    def _update_reading(self):
        try:
            p = self._get_current_params()
            if self.experiment_type == "distance":
                voc = solve_voc(p)
                isc = solve_isc(p)
                self.reading_v.config(text="Voc = {:.3f} V".format(voc))
                self.reading_i.config(text="Isc = {:.3f} mA".format(isc * 1000))
                self.reading_p.config(text="E = {:.0f} W/m²".format(self.light_intensity))
            else:
                V, I, P = solve_operating_point(p, self.fixed_R)
                self.reading_v.config(text="U = {:.3f} V".format(V))
                self.reading_i.config(text="I = {:.3f} mA".format(I * 1000))
                self.reading_p.config(text="P = {:.3f} mW".format(P * 1000))
        except Exception:
            self.reading_v.config(text="U = --- V")
            self.reading_i.config(text="I = --- mA")
            self.reading_p.config(text="P = --- mW")

    def _record_point(self):
        if self.experiment_type == "distance":
            var_val = self.distance_cm
            var_name = "d"
        else:
            var_val = self.light_intensity
            var_name = "E"

        for dp in self.data_points:
            if abs(dp["var_val"] - var_val) < 1e-6:
                self._show_toast("已存在 {}={:g} 的数据，不可重复记录！".format(var_name, var_val))
                return

        try:
            p = self._get_current_params()
            if self.experiment_type == "distance":
                voc = solve_voc(p)
                isc = solve_isc(p)
                point = {
                    "var_name": var_name, "var_val": var_val,
                    "E": self.light_intensity, "Voc": voc, "Isc": isc
                }
            else:
                V, I, P = solve_operating_point(p, self.fixed_R)
                point = {
                    "var_name": var_name, "var_val": var_val,
                    "E": self.light_intensity, "V": V, "I": I, "P": P
                }
        except Exception:
            return

        self.data_points.append(point)

        idx = len(self.data_points)
        if self.experiment_type == "distance":
            self.tree.insert("", "end", values=(
                idx, "{:.0f}".format(var_val), "{:.0f}".format(self.light_intensity),
                "{:.3f}".format(point["Voc"]), "{:.3f}".format(point["Isc"] * 1000)
            ))
        else:
            self.tree.insert("", "end", values=(
                idx, "{:.0f}".format(var_val), "{:.3f}".format(point["V"]),
                "{:.3f}".format(point["I"] * 1000), "{:.3f}".format(point["P"] * 1000)
            ))
        self._draw_plot()

    def _delete_selected(self):
        sel = self.tree.selection()
        if not sel:
            self._show_toast("请先在表格中选中要删除的行")
            return
        for item_id in sel:
            vals = self.tree.item(item_id, "values")
            var_val = float(vals[1])
            self.data_points = [dp for dp in self.data_points
                                if abs(dp["var_val"] - var_val) > 1e-6]
            self.tree.delete(item_id)
        self._reindex_table()
        self._draw_plot()

    def _reindex_table(self):
        for i, item_id in enumerate(self.tree.get_children()):
            vals = list(self.tree.item(item_id, "values"))
            vals[0] = i + 1
            self.tree.item(item_id, values=vals)

    def _show_toast(self, msg):
        toast = tk.Toplevel(self.frame.winfo_toplevel())
        toast.overrideredirect(True)
        toast.configure(bg="#ff4444")
        x = self.frame.winfo_rootx() + self.frame.winfo_width() // 2 - 150
        y = self.frame.winfo_rooty() + 80
        toast.geometry("300x36+{}+{}".format(x, y))
        tk.Label(toast, text=msg, bg="#ff4444", fg="#fff",
                 font=("Microsoft YaHei", 10, "bold")).pack(expand=True, fill=tk.BOTH)
        toast.after(2000, toast.destroy)

    def _clear_data(self):
        self.data_points.clear()
        for item_id in self.tree.get_children():
            self.tree.delete(item_id)
        if self.ax2 is not None:
            self.ax2.remove()
            self.ax2 = None
        self.ax.clear()
        self._style_ax(self.ax)
        self.canvas.draw_idle()

    def _draw_plot(self):
        if not self.data_points:
            if self.ax2 is not None:
                self.ax2.remove()
                self.ax2 = None
            self.ax.clear()
            self._style_ax(self.ax)
            self.canvas.draw_idle()
            return
        if self.ax2 is not None:
            self.ax2.remove()
            self.ax2 = None
        self.ax.clear()
        var_name = self.data_points[0]["var_name"]
        if self.experiment_type == "distance":
            pts = sorted(self.data_points, key=lambda d: d["E"])
            xs = [d["E"] for d in pts]
            voc = [d["Voc"] for d in pts]
            isc = [d["Isc"] * 1000 for d in pts]

            self.ax.set_xlabel("E (W/m²)", fontproperties=plt_font, color=FG, fontsize=10)
            self.ax.set_ylabel("Voc (V)", fontproperties=plt_font, color=ACCENT, fontsize=10)
            self.ax.tick_params(axis="y", labelcolor=ACCENT)
            self.ax.plot(xs, voc, "o-", color=ACCENT, linewidth=2,
                         markersize=6, label="Voc (V)")

            ax2 = self.ax.twinx()
            self.ax2 = ax2
            ax2.plot(xs, isc, "s--", color=ACCENT2, linewidth=2,
                     markersize=6, label="Isc (mA)")
            ax2.set_ylabel("Isc (mA)", fontproperties=plt_font, color=ACCENT2, fontsize=10, labelpad=10)
            ax2.tick_params(axis="y", labelcolor=ACCENT2)
            for spine in ax2.spines.values():
                spine.set_color(GRID_COLOR)
            if isc:
                ax2.set_ylim(0, max(isc) * 1.25)
        else:
            pts = sorted(self.data_points, key=lambda d: d["var_val"])
            xs = [d["var_val"] for d in pts]
            current = [d["I"] * 1000 for d in pts]
            power = [d["P"] * 1000 for d in pts]

            self.ax.set_xlabel(var_name, fontproperties=plt_font, color=FG, fontsize=10)
            self.ax.set_ylabel("I (mA)", fontproperties=plt_font, color=ACCENT, fontsize=10)
            self.ax.tick_params(axis="y", labelcolor=ACCENT)
            self.ax.plot(xs, current, "o-", color=ACCENT, linewidth=2,
                         markersize=6, label="I (mA)")

            ax2 = self.ax.twinx()
            self.ax2 = ax2
            ax2.plot(xs, power, "s--", color=ACCENT2, linewidth=2,
                     markersize=6, label="P (mW)")
            ax2.set_ylabel("P (mW)", fontproperties=plt_font, color=ACCENT2, fontsize=10, labelpad=10)
            ax2.tick_params(axis="y", labelcolor=ACCENT2)
            for spine in ax2.spines.values():
                spine.set_color(GRID_COLOR)
            ax2.set_ylim(0, max(max(power), 1) * 1.25)

        self.ax.set_title("实验数据 - " + self.title, fontproperties=plt_font,
                          color=FG, fontsize=12, pad=8)
        lines1, labels1 = self.ax.get_legend_handles_labels()
        lines2, labels2 = ax2.get_legend_handles_labels()
        self.ax.legend(lines1 + lines2, labels1 + labels2,
                       loc="upper center", bbox_to_anchor=(0.5, 1.02),
                       ncol=2, fontsize=9, framealpha=0.4,
                       labelcolor=FG, prop=plt_font)
        self.fig.tight_layout()
        self.canvas.draw_idle()


# ═══════════════════════════════════════════════
#  扩展：15块电池驱动LED优化
# ═══════════════════════════════════════════════

class EssayTab:
    """扩展：给定 N，求 (m, n, x) 最优组合。"""

    def __init__(self, parent_frame):
        self.frame = parent_frame
        self._build()

    def _build(self):
        wrap = tk.Frame(self.frame, bg=BG)
        wrap.pack(fill=tk.BOTH, expand=True, padx=0, pady=0)

        left = tk.Frame(wrap, bg=PANEL_BG, width=160)
        left.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 10))
        left.pack_propagate(False)
        right = tk.Frame(wrap, bg=BG)
        right.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        self.vars = {
            "N": tk.DoubleVar(value=15),
            "Vm": tk.DoubleVar(value=2.16),
            "Im": tk.DoubleVar(value=21.7),
            "E": tk.DoubleVar(value=242.0),
            "Vled": tk.DoubleVar(value=2.8),
            "Iled": tk.DoubleVar(value=8.0),
        }

        fields = [
            ("电池数量 N", "N"),
            ("单块 Vm (V)", "Vm"),
            ("单块 Im (mA)", "Im"),
            ("光强 E (W/m²)", "E"),
            ("LED 额定电压 (V)", "Vled"),
            ("LED 额定电流 (mA)", "Iled"),
        ]
        readonly_keys = {"Vm", "Im", "E"}
        for text, key in fields:
            row = tk.Frame(left, bg=PANEL_BG)
            row.pack(fill=tk.X, padx=10, pady=3)
            tk.Label(row, text=text, bg=PANEL_BG, fg=FG, anchor="w",
                     font=("Microsoft YaHei", 10)).pack(fill=tk.X)
            ent = tk.Entry(row, textvariable=self.vars[key], font=("Consolas", 11),
                           bg="#fff", fg="#111")
            if key in readonly_keys:
                ent.config(state="disabled", disabledbackground="#d9d9d9", disabledforeground="#666")
            ent.pack(fill=tk.X, pady=(2, 0))

        ttk.Button(left, text="计算方案", style="Info.TButton",
                   command=self._calc).pack(fill=tk.X, padx=10, pady=(10, 8), ipady=2)

        table_wrap = tk.Frame(right, bg=BG)
        table_wrap.pack(fill=tk.BOTH, expand=True, pady=(0, 4))
        cols = ("序号", "每组并联电池数m", "串联组数n", "可点亮LED总数x", "输出电压Vout(V)", "输出电流Iout(mA)", "每支路串联LED数", "可并联支路数")
        self.result_tree = ttk.Treeview(table_wrap, columns=cols, show="headings", height=14)
        widths = (60, 170, 120, 170, 150, 170, 170, 140)
        for c, w in zip(cols, widths):
            self.result_tree.heading(c, text=c)
            self.result_tree.column(c, width=w, minwidth=80, anchor="center", stretch=True)
        _style = ttk.Style()
        _style.configure(
            "Expand.Treeview",
            background=BG,
            fieldbackground=BG,
            bordercolor=BG,
            lightcolor=BG,
            darkcolor=BG
        )
        _style.configure("Expand.Treeview.Heading", background=PANEL_BG, foreground=FG)
        self.result_tree.configure(style="Expand.Treeview")
        self.result_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=0, pady=0)
        self.result_tree.tag_configure("best", background="#ffe56a", foreground="#111111")

        self.summary_lbl = tk.Label(right, text="", bg=BG, fg=FG, justify="left",
                                    anchor="w", font=("Microsoft YaHei", 10))
        self.summary_lbl.pack(fill=tk.X, padx=4, pady=(2, 0))
        self._calc()

    def _calc(self):
        try:
            N = int(self.vars["N"].get())
            Vm = float(self.vars["Vm"].get())
            Im_mA = float(self.vars["Im"].get())
            Vled = float(self.vars["Vled"].get())
            Iled_mA = float(self.vars["Iled"].get())
            if min(N, Vm, Im_mA, Vled, Iled_mA) <= 0:
                raise ValueError
        except Exception:
            for item in self.result_tree.get_children():
                self.result_tree.delete(item)
            self.summary_lbl.config(text="输入参数无效，请输入大于0的数值。")
            return

        Im = Im_mA / 1000.0
        Iled = Iled_mA / 1000.0

        combos = []
        for m in range(1, N + 1):
            if N % m != 0:
                continue
            n = N // m
            v_out = n * Vm
            i_out = m * Im
            led_per_branch = int(v_out // Vled)
            branch_count = int(i_out // Iled)
            x = led_per_branch * branch_count
            combos.append({
                "m": m, "n": n, "x": x,
                "v": v_out, "i_mA": i_out * 1000.0,
                "led_per_branch": led_per_branch, "branch_count": branch_count
            })
        best = None
        if combos:
            best = max(combos, key=lambda c: (c["x"], -c["m"]))
        self.last_combos = combos
        self.best_combo = best

        for item in self.result_tree.get_children():
            self.result_tree.delete(item)
        for idx, c in enumerate(combos, start=1):
            tags = ("best",) if best is not None and c["m"] == best["m"] and c["n"] == best["n"] else ()
            self.result_tree.insert("", "end", values=(
                idx, c["m"], c["n"], c["x"],
                f"{c['v']:.2f}", f"{c['i_mA']:.2f}",
                c["led_per_branch"], c["branch_count"]
            ), tags=tags)
        if not combos:
            self.summary_lbl.config(text="无可行组合。")
            return
        self.summary_lbl.config(
            text=(
                f"输入: N={N}, Vm={Vm:.2f}V, Im={Im_mA:.1f}mA, Vled={Vled:.2f}V, Iled={Iled_mA:.1f}mA\n"
                f"最优组合: m={best['m']}, n={best['n']}, x={best['x']}（理想整除估算）"
            )
        )

    def export_data_excel(self):
        combos = getattr(self, "last_combos", [])
        if not combos:
            messagebox.showinfo("提示", "扩展页暂无可导出数据")
            return
        path = filedialog.asksaveasfilename(
            title="导出扩展页数据",
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx")],
            initialfile="扩展_LED优化_数据.xlsx",
        )
        if not path:
            return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill
        except Exception:
            messagebox.showerror("错误", "缺少 openpyxl，请先安装：pip install openpyxl")
            return
        wb = Workbook()
        ws = wb.active
        ws.title = "组合结果"
        headers = ["序号", "每组并联电池数m", "串联组数n", "可点亮LED总数x", "输出电压Vout(V)", "输出电流Iout(mA)", "每支路串联LED数", "可并联支路数"]
        ws.append(headers)
        best = getattr(self, "best_combo", None)
        best_fill = PatternFill(fill_type="solid", start_color="FFF2CC", end_color="FFF2CC")
        for idx, c in enumerate(combos, start=1):
            ws.append([idx, c["m"], c["n"], c["x"], c["v"], c["i_mA"], c["led_per_branch"], c["branch_count"]])
            if best and c["m"] == best["m"] and c["n"] == best["n"] and c["x"] == best["x"]:
                for col in range(1, 9):
                    ws.cell(row=idx + 1, column=col).fill = best_fill
        widths = [8, 18, 12, 16, 16, 16, 16, 14]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + i)].width = w
        wb.save(path)
        messagebox.showinfo("完成", f"已导出：\n{path}")

    def export_chart(self):
        messagebox.showinfo("提示", "扩展页暂无图表可导出")


# ═══════════════════════════════════════════════
#  主应用
# ═══════════════════════════════════════════════

class SolarCellApp:
    def __init__(self, root):
        self.root = root
        self.root.title("太阳能电池特性测量")
        self.root.geometry("1820x900")
        self.root.configure(bg=BG)
        self.root.minsize(1560, 760)
        configure_ttk_theme(self.root)
        self._build_ui()

    def _build_ui(self):
        header = tk.Frame(self.root, bg=ACCENT, height=36)
        header.pack(fill=tk.X)
        header.pack_propagate(False)
        tk.Label(header, text="太阳能电池的特性测量",
                 bg=ACCENT, fg="#fff",
                 font=("Microsoft YaHei", 13, "bold")).pack(side=tk.LEFT, padx=12)
        right = tk.Frame(header, bg=ACCENT)
        right.pack(side=tk.RIGHT, padx=8)
        ttk.Button(right, text="工具箱", style="Assist.TButton",
                   command=self._open_toolbox).pack(side=tk.LEFT, padx=(0, 4))
        ttk.Button(right, text="帮助", style="Assist.TButton",
                   command=self._open_help).pack(side=tk.LEFT)

        main = tk.Frame(self.root, bg=BG)
        main.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)
        self.tabs = ttk.Notebook(main)
        self.tabs.pack(fill=tk.BOTH, expand=True)

        exp1_frame = tk.Frame(self.tabs, bg=BG)
        exp2_frame = tk.Frame(self.tabs, bg=BG)
        essay_frame = tk.Frame(self.tabs, bg=BG)
        self.tabs.add(exp1_frame, text="实验一  伏安特性")
        self.tabs.add(exp2_frame, text="实验二  距离特性")
        self.tabs.add(essay_frame, text="扩展  LED优化")

        self.exp1 = ExperimentOneTab(exp1_frame, experiment_name="实验一")
        self.exp2 = ExperimentOneTab(exp2_frame, experiment_name="实验二")
        self.essay = EssayTab(essay_frame)

    def _current_tab_key(self):
        idx = self.tabs.index(self.tabs.select())
        if idx == 0:
            return "exp1"
        if idx == 1:
            return "exp2"
        return "ext"

    def _open_toolbox(self):
        menu = tk.Menu(self.root, tearoff=0, bg="#f7f7f7", fg="#111111",
                       activebackground="#d9ecff", activeforeground="#000000")
        key = self._current_tab_key()
        if key == "exp1":
            menu.add_command(label="一键连线", command=self.exp1._auto_wire_for_test)
            menu.add_command(label="一键导入实验一标准数据", command=self.exp1._load_standard_data)
            menu.add_separator()
            menu.add_command(label="导出本实验数据（Excel）", command=self.exp1.export_data_excel)
            menu.add_command(label="导出本实验图表", command=self.exp1.export_chart)
        elif key == "exp2":
            menu.add_command(label="一键连线 Isc", command=self.exp2._auto_wire_exp2_isc)
            menu.add_command(label="一键连线 Voc", command=self.exp2._auto_wire_exp2_voc)
            menu.add_command(label="一键导入实验二标准数据", command=self.exp2._load_standard_data)
            menu.add_separator()
            menu.add_command(label="导出本实验数据（Excel）", command=self.exp2.export_data_excel)
            menu.add_command(label="导出本实验图表", command=self.exp2.export_chart)
        else:
            menu.add_command(label="导出本实验数据（Excel）", command=self.essay.export_data_excel)
            menu.add_command(label="导出本实验图表", command=self.essay.export_chart)
        try:
            x, y = self.root.winfo_pointerxy()
            menu.tk_popup(x, y)
        finally:
            menu.grab_release()

    def _open_help(self):
        win = tk.Toplevel(self.root)
        win.title("帮助")
        win.configure(bg=BG)
        win.geometry("460x260")
        win.resizable(False, False)

        tk.Label(win, text="使用说明", bg=ACCENT, fg="#fff",
                 font=("Microsoft YaHei", 12, "bold")).pack(fill=tk.X, ipady=6)
        key = self._current_tab_key()
        if key == "exp1":
            msg = (
                "实验一帮助：\n"
                "1. 完成接线后记录数据（或自动填充标准数据）。\n"
                "2. 打开特性曲线窗口查看 I-V / P-V。\n"
                "3. 曲线窗口可切换拟合曲线与原始图像。"
            )
        elif key == "exp2":
            msg = (
                "实验二帮助：\n"
                "1. 在 Isc/Voc 两种模式下分别接线并记录。\n"
                "2. 同一距离下需要完成两次测量后形成完整数据。\n"
                "3. 打开曲线窗口查看 Voc-I 与 Isc-I。"
            )
        else:
            msg = (
                "扩展页帮助：\n"
                "1. 输入 N、Vm、Im、Vled、Iled 参数。\n"
                "2. 点击“计算方案”得到全部 (m, n, x) 组合。\n"
                "3. 黄色高亮行为当前最优组合。"
            )
        tk.Label(win, text=msg, bg=BG, fg=FG, justify="left",
                 anchor="nw", font=("Microsoft YaHei", 10)).pack(fill=tk.BOTH, expand=True, padx=12, pady=12)


def run():
    root = tk.Tk()
    app = SolarCellApp(root)
    root.mainloop()


if __name__ == "__main__":
    run()
